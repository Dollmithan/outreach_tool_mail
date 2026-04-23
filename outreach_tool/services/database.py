"""
Database layer — all persistence via Supabase REST API.
No local SQLite files; no working copies.
"""
import os
import re
from datetime import datetime, timezone
from typing import List, Optional, Tuple

from .supabase_client import (
    sb_batch_insert, sb_count, sb_delete, sb_insert,
    sb_select, sb_update, sb_upsert,
)


def normalize_email(addr):
    return (addr or "").strip().lower()


# ─── Imports ────────────────────────────────────────────────────────────────────

def list_imports():
    return sb_select("imports", order="id.desc")


def get_stats_for_import(import_id):
    total = sb_count("leads", filters={"import_id": import_id})
    sent = sb_count("leads", filters={"import_id": import_id, "sent": True})
    replied = sb_count("leads", filters={"import_id": import_id, "replied": True})
    left = total - sent
    return {
        "total": total,
        "sent": sent,
        "replied": replied,
        "left": left,
        "total_rows": total,
        "no_email": 0,
    }


def get_numbers_for_import(import_id):
    return sb_select("leads", filters={"import_id": import_id}, columns="number")


def remove_import(import_id):
    sb_delete("leads", filters={"import_id": import_id})
    sb_delete("imports", filters={"id": import_id})


def add_lead_to_import(import_id, full_name, number, raw_email):
    e = normalize_email(raw_email)
    if not e:
        raise ValueError("Email is required.")
    sb_upsert("leads", {
        "import_id": import_id,
        "email": e,
        "full_name": full_name or "",
        "number": number or "",
        "sent": False,
        "replied": False,
    })


def resync_import_from_working(import_id):
    pass  # No-op: no working copies in Supabase mode


# ─── Leads ──────────────────────────────────────────────────────────────────────

def get_unsent(import_id) -> List[Tuple[str, str, str]]:
    rows = sb_select("leads", filters={"import_id": import_id, "sent": False})
    return [(r.get("full_name", ""), r.get("number", ""), r["email"]) for r in rows]


def mark_sent(import_id, email_addr):
    e = normalize_email(email_addr)
    ts = datetime.now(timezone.utc).isoformat()
    sb_update("leads", {"sent": True, "sent_at": ts},
              filters={"import_id": import_id, "email": e})


def mark_replied_everywhere(email_addr):
    e = normalize_email(email_addr)
    ts = datetime.now(timezone.utc).isoformat()
    sb_update("leads", {"replied": True, "replied_at": ts},
              extra={"email": f"eq.{e}", "sent": "eq.true"})


def lookup_contact_master(email_addr):
    e = normalize_email(email_addr)
    rows = sb_select("leads", extra={"email": f"eq.{e}", "sent": "eq.true"},
                     order="id.desc", limit=1)
    if not rows:
        rows = sb_select("leads", extra={"email": f"eq.{e}"},
                         order="id.desc", limit=1)
    if not rows:
        return None
    r = rows[0]
    return (r.get("full_name", ""), r.get("number", ""), r["email"], r.get("sent_at"))


def is_reply_processed(account_key, message_id):
    rows = sb_select("processed_replies",
                     filters={"account_key": str(account_key),
                               "message_id": str(message_id)})
    return bool(rows)


def record_reply_processed(account_key, message_id):
    try:
        sb_insert("processed_replies", {
            "account_key": str(account_key),
            "message_id": str(message_id),
        })
    except Exception:
        pass


# ─── Importing data ──────────────────────────────────────────────────────────────

def _batch_push_leads(import_id, rows):
    """rows: iterable of (full_name, number, email, sent, replied, replied_at, sent_at)"""
    batch = []
    seen_emails = set()
    for row in rows:
        full_name, number, email, sent, replied, replied_at, sent_at = row
        e = normalize_email(email)
        if not e or e in seen_emails:
            continue
        seen_emails.add(e)
        batch.append({
            "import_id": import_id,
            "email": e,
            "full_name": full_name or "",
            "number": number or "",
            "sent": bool(sent),
            "replied": bool(replied),
            "sent_at": sent_at or None,
            "replied_at": replied_at or None,
        })
    if batch:
        sb_batch_insert("leads", batch)
    return len(batch)


def import_user_database(sqlite_path, label=None):
    """Read a .db SQLite file and push contacts into Supabase leads. Returns (import_id, None)."""
    import sqlite3
    label = label or "Imported DB"
    result = sb_insert("imports", {"label": label})
    import_id = result[0]["id"]
    conn = sqlite3.connect(sqlite_path)
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(contacts)")
    cols = {r[1] for r in cur.fetchall()}
    sent_col = "COALESCE(sent, 0)" if "sent" in cols else "0"
    replied_col = "COALESCE(replied, 0)" if "replied" in cols else "0"
    replied_at_col = "replied_at" if "replied_at" in cols else "NULL"
    sent_at_col = "sent_at" if "sent_at" in cols else "NULL"
    cur.execute(f"""
        SELECT full_name, number, email,
               {sent_col}, {replied_col},
               {replied_at_col}, {sent_at_col}
        FROM contacts
        WHERE email IS NOT NULL AND TRIM(COALESCE(email, '')) != ''
    """)
    rows = cur.fetchall()
    conn.close()
    _batch_push_leads(import_id, rows)
    return import_id, None


def import_excel_as_leads(excel_path, label=None):
    """Parse Excel and push leads into Supabase. Returns (import_id, None)."""
    rows = _parse_excel_to_rows(excel_path)
    if not rows:
        raise ValueError(
            "No rows with an email address were imported. "
            "Check the header row and that data starts on row 2."
        )
    label = label or "Imported Excel"
    result = sb_insert("imports", {"label": label})
    import_id = result[0]["id"]
    _batch_push_leads(import_id, rows)
    return import_id, None


# ─── Excel parsing (returns list of row tuples, no SQLite) ──────────────────────

_EMAIL_RE = re.compile(r"^[A-Z0-9._%+\-']+@[A-Z0-9.\-]+\.[A-Z]{2,}$", re.IGNORECASE)


def _excel_header_norm(cell):
    if cell is None:
        return ""
    return re.sub(r"\s+", " ", str(cell).strip().lower())


def _is_email_like(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    if not s or " " in s or "@" not in s:
        return False
    return _EMAIL_RE.match(s) is not None


def _digits_only(val) -> str:
    if val is None:
        return ""
    return re.sub(r"\D+", "", str(val))


def _is_phone_like(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    if not s or "@" in s:
        return False
    d = _digits_only(s)
    return len(d) >= 7


def _excel_cell_to_int01(val, default=0):
    if val is None or val == "":
        return default
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, (int, float)):
        return 1 if int(val) != 0 else 0
    s = str(val).strip().lower()
    if s in ("1", "true", "yes", "y", "x"):
        return 1
    if s in ("0", "false", "no", "n"):
        return 0
    return default


def _score_excel_columns(header_vals, sample_rows, max_cols: int):
    headers = [_excel_header_norm(c) for c in (header_vals or ())]
    idx = {
        "full_name": None,
        "first_name": None,
        "last_name": None,
        "email": None,
        "number": None,
        "sent": None,
        "replied": None,
    }

    for i, h in enumerate(headers[:max_cols]):
        if not h:
            continue
        if idx["sent"] is None and h == "sent":
            idx["sent"] = i
        if idx["replied"] is None and h in ("replied", "reply"):
            idx["replied"] = i
        if idx["first_name"] is None and (
            (("first" in h) and ("name" in h))
            or h in ("firstname", "given name", "givenname", "given")
        ):
            idx["first_name"] = i
        if idx["last_name"] is None and (
            (("last" in h) and ("name" in h))
            or h in ("lastname", "surname", "family name", "familyname", "last")
        ):
            idx["last_name"] = i
        if idx["full_name"] is None and (("full" in h) and ("name" in h)):
            idx["full_name"] = i

    email_scores = [0.0] * max_cols
    phone_scores = [0.0] * max_cols
    name_scores = [0.0] * max_cols

    for i in range(max_cols):
        h = headers[i] if i < len(headers) else ""
        if any(x in h for x in ("email", "e-mail", "e mail", "mail")):
            email_scores[i] += 3.0
        if any(x in h for x in ("phone", "mobile", "tel", "telephone", "cell", "whatsapp")) or h == "number":
            phone_scores[i] += 3.0
        if any(x in h for x in ("full name", "fullname", "contact name", "client name", "lead name")) or h in ("name", "contact"):
            name_scores[i] += 2.0
        if idx.get("first_name") == i:
            name_scores[i] += 1.5
        if idx.get("last_name") == i:
            name_scores[i] += 1.5

    for row in sample_rows or ():
        if not row:
            continue
        for i in range(max_cols):
            val = row[i] if i < len(row) else None
            if _is_email_like(val):
                email_scores[i] += 1.0
            if _is_phone_like(val):
                phone_scores[i] += 1.0
            if val is not None:
                s = str(val).strip()
                if s and not _is_email_like(s):
                    d = _digits_only(s)
                    if len(d) <= 2 and len(s) >= 3:
                        name_scores[i] += 0.2

    def best_index(scores):
        best_i, best_v = None, -1e9
        for i, v in enumerate(scores):
            if v > best_v:
                best_i, best_v = i, v
        return best_i, best_v

    email_i, email_v = best_index(email_scores)
    phone_i, phone_v = best_index(phone_scores)

    if email_i is not None and phone_i is not None and email_i == phone_i:
        if email_v >= phone_v:
            phone_scores[email_i] = -1e9
            phone_i, phone_v = best_index(phone_scores)
        else:
            email_scores[phone_i] = -1e9
            email_i, email_v = best_index(email_scores)

    idx["email"] = email_i if (email_i is not None and email_v >= 2.0) else None
    idx["number"] = phone_i if (phone_i is not None and phone_v >= 2.0) else None

    forbidden = {idx["email"], idx["number"], idx["sent"], idx["replied"]}
    best_name_i, best_name_v = None, -1e9
    for i, v in enumerate(name_scores):
        if i in forbidden:
            continue
        if v > best_name_v:
            best_name_i, best_name_v = i, v
    if best_name_i is not None and best_name_v > 0.5:
        idx["full_name"] = best_name_i
    else:
        for i in range(max_cols):
            if i not in forbidden:
                idx["full_name"] = i
                break

    return idx


def _map_excel_columns(header_vals, sample_rows):
    max_cols = max(len(header_vals or ()), max((len(r) for r in (sample_rows or ())), default=0), 0)
    max_cols = min(max_cols, 80)
    idx = _score_excel_columns(header_vals or (), sample_rows or (), max_cols)

    if idx["email"] is None and max_cols >= 3:
        col1_em = 0
        for r in (sample_rows or ())[:20]:
            if len(r) > 1 and _is_email_like(r[1]):
                col1_em += 1
        if col1_em >= 3:
            idx["full_name"], idx["email"], idx["number"] = 0, 1, 2
            if max_cols >= 5:
                idx["sent"], idx["replied"] = 3, 4
            elif max_cols >= 4:
                idx["sent"] = 3

    return idx


def _parse_excel_to_rows(excel_path):
    """Parse an Excel file and return list of (full_name, number, email, sent, replied, replied_at, sent_at)."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl is required for Excel import.") from e

    import os
    ext = os.path.splitext(excel_path)[1].lower()
    if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raise ValueError("Supported Excel types: .xlsx / .xlsm")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    best = None

    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_vals = next(rows_iter)
        except StopIteration:
            continue

        sample_rows = []
        for _ in range(50):
            try:
                r = next(rows_iter)
            except StopIteration:
                break
            sample_rows.append(r)

        col = _map_excel_columns(header_vals, sample_rows)
        if col.get("email") is None:
            continue

        email_hits = sum(
            1 for r in sample_rows
            if r and col.get("email") is not None
            and col["email"] < len(r) and _is_email_like(r[col["email"]])
        )
        phone_hits = sum(
            1 for r in sample_rows
            if r and col.get("number") is not None
            and col["number"] < len(r) and _is_phone_like(r[col["number"]])
        )
        score = (email_hits * 10.0) + (phone_hits * 1.0)
        if best is None or score > best[0]:
            best = (score, ws, header_vals, sample_rows, col)

    if best is None:
        try:
            wb.close()
        except Exception:
            pass
        raise ValueError("Could not find an Email column in any sheet.")

    _, ws, header_vals, sample_rows, col = best
    output = []

    def process_row(row):
        if row is None:
            return
        def gv(key):
            ix = col[key]
            if ix is None or ix >= len(row):
                return None
            return row[ix]

        email_raw = gv("email")
        em = normalize_email(email_raw or "")
        if not em:
            return
        first = gv("first_name")
        last = gv("last_name")
        if (first is not None and str(first).strip()) or (last is not None and str(last).strip()):
            first_s = str(first).strip() if first is not None else ""
            last_s = str(last).strip() if last is not None else ""
            full_name = (first_s + (" " if (first_s and last_s) else "") + last_s).strip()
        else:
            fn = gv("full_name")
            full_name = (str(fn).strip() if fn is not None else "") or ""
        num = gv("number")
        number = (str(num).strip() if num is not None else "") or ""
        sent = _excel_cell_to_int01(gv("sent"), 0)
        replied = _excel_cell_to_int01(gv("replied"), 0)
        output.append((full_name, number, em, sent, replied, None, None))

    for row in sample_rows:
        process_row(row)

    rows_iter = ws.iter_rows(values_only=True)
    try:
        _ = next(rows_iter)
    except StopIteration:
        rows_iter = iter(())
    for _ in range(len(sample_rows)):
        try:
            _ = next(rows_iter)
        except StopIteration:
            break
    for row in rows_iter:
        process_row(row)

    try:
        wb.close()
    except Exception:
        pass

    return output
