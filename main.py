import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import sqlite3
import smtplib
import imaplib
import email
import email.utils
import json
import time
import random
import os
import re
import ssl
import shutil
import uuid
import requests
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import configparser
import logging
import phonenumbers
import dns.resolver
from typing import List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

CONFIG_FILE_NAME = "config.ini"

APP_DIR = os.path.dirname(os.path.abspath(__file__))
APP_DATA_NAME = "outreach_tool_data"
_APP_DATA_DIR = None

logger = logging.getLogger(__name__)

_VERIFY_EMAIL_SYNTAX_RE = re.compile(
    r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$",
    re.IGNORECASE,
)


def _verification_log(msg: str, log_fn) -> None:
    logger.info(msg)
    if log_fn:
        log_fn(msg)


def check_email_syntax(email: str) -> bool:
    s = (email or "").strip()
    if not s or " " in s or s.count("@") != 1:
        return False
    return _VERIFY_EMAIL_SYNTAX_RE.match(s) is not None


def list_mx_hosts(domain: str) -> List[str]:
    try:
        records = dns.resolver.resolve(domain, "MX")
        ordered = sorted(records, key=lambda r: (r.preference, str(r.exchange)))
        return [str(r.exchange).rstrip(".") for r in ordered]
    except Exception:
        return []


def _helo_domain(mail_from: str) -> str:
    return mail_from.rsplit("@", 1)[-1] if "@" in mail_from else "localhost"


def _rcpt_code_verdict(code: int) -> Optional[bool]:
    """True = accepted, False = permanent reject, None = inconclusive (e.g. greylist 4xx)."""
    if 200 <= code < 300:
        return True
    if 500 <= code < 600:
        return False
    return None


def _smtp_handshake_rcpt(
    smtp: smtplib.SMTP, email: str, mail_from: str
) -> Optional[bool]:
    helo = _helo_domain(mail_from)
    try:
        smtp.ehlo(helo)
    except smtplib.SMTPException:
        try:
            smtp.helo(helo)
        except smtplib.SMTPException:
            return None
    try:
        smtp.mail(mail_from)
    except smtplib.SMTPException:
        return None
    try:
        code, _ = smtp.rcpt(email)
        return _rcpt_code_verdict(code)
    except smtplib.SMTPRecipientsRefused:
        return False


def _mx_probe_attempt(email: str, mail_from: str, connect) -> Optional[bool]:
    smtp = None
    try:
        smtp = connect()
        return _smtp_handshake_rcpt(smtp, email, mail_from)
    except Exception:
        return None
    finally:
        if smtp is not None:
            try:
                smtp.quit()
            except Exception:
                try:
                    smtp.close()
                except Exception:
                    pass


def probe_rcpt_on_mx_multiport(
    email: str, mx_host: str, mail_from: str
) -> Optional[bool]:
    """
    Try several ports / TLS modes on the recipient MX (ISP often blocks only :25).
    True / False / None same as _smtp_handshake_rcpt; None if no strategy connected.
    """
    ctx = ssl.create_default_context()
    timeout = 12

    def plain25():
        s = smtplib.SMTP(timeout=timeout)
        s.connect(mx_host, 25)
        return s

    def starttls587():
        s = smtplib.SMTP(timeout=timeout)
        s.connect(mx_host, 587)
        s.ehlo()
        s.starttls(context=ctx)
        s.ehlo()
        return s

    def ssl465():
        return smtplib.SMTP_SSL(mx_host, 465, timeout=timeout, context=ctx)

    for label, factory in (
        ("25", plain25),
        ("587+STARTTLS", starttls587),
        ("465 SSL", ssl465),
    ):
        result = _mx_probe_attempt(email, mail_from, factory)
        if result is not None:
            return result
    return None


def verify_email(
    addr: str, log_fn=None, mail_from: Optional[str] = None
) -> Tuple[bool, str]:
    """
    Pre-send: syntax, MX list, RCPT probe per MX (ports 25 / 587 / 465), then RCPT via
    your configured SMTP if direct MX is unreachable (e.g. outbound port 25 blocked).
    Inconclusive SMTP result is treated as valid (we still attempt delivery).
    Returns (ok_to_send, reason_for_failure_if_any).
    """
    email_raw = (addr or "").strip()
    _verification_log(f"📧 Verifying {email_raw!r} before send...", log_fn)

    if not check_email_syntax(email_raw):
        _verification_log("   ✗ Syntax check failed", log_fn)
        return False, "invalid email syntax"

    domain = email_raw.split("@", 1)[1]
    _verification_log(f"   Syntax OK — MX lookup for {domain!r}...", log_fn)
    mx_hosts = list_mx_hosts(domain)
    if not mx_hosts:
        _verification_log("   ✗ No MX records (or lookup failed)", log_fn)
        return False, "no MX records for recipient domain"

    from_addr = (mail_from or "").strip() or "verify@localhost"
    n = len(mx_hosts)
    for i, mx in enumerate(mx_hosts):
        _verification_log(
            f"   MX [{i + 1}/{n}] {mx} — RCPT probe (25 → 587+TLS → 465)...",
            log_fn,
        )
        smtp_result = probe_rcpt_on_mx_multiport(email_raw, mx, from_addr)
        if smtp_result is True:
            _verification_log("   ✓ SMTP RCPT accepted (2xx)", log_fn)
            return True, "ok"
        if smtp_result is False:
            _verification_log("   ✗ SMTP RCPT rejected — skipping send", log_fn)
            return False, "recipient rejected by recipient mail server (RCPT)"

    _verification_log(
        "   Direct MX not reachable or non-committal — RCPT via your SMTP (configured host)...",
        log_fn,
    )
    relay_result = probe_rcpt_via_configured_smtp(email_raw, from_addr, log_fn)
    if relay_result is True:
        _verification_log("   ✓ Relay SMTP accepted RCPT (2xx)", log_fn)
        return True, "ok"
    if relay_result is False:
        _verification_log("   ✗ Relay SMTP rejected RCPT — skipping send", log_fn)
        return False, "recipient rejected by your outbound mail server (RCPT)"

    _verification_log(
        "   ? Verification inconclusive — keeping recipient (will try outbound SMTP)",
        log_fn,
    )
    return True, "ok"

def _resolve_app_data_dir():
    """
    Pick a writable app data location.
    1) Prefer project-local folder for portability.
    2) Fallback to %LOCALAPPDATA% on Windows if project folder is not writable.
    """
    local_candidate = os.path.join(APP_DIR, "app_data")
    fallback_base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    fallback_candidate = os.path.join(fallback_base, APP_DATA_NAME)
    for candidate in (local_candidate, fallback_candidate):
        try:
            os.makedirs(candidate, exist_ok=True)
            probe = os.path.join(candidate, ".write_test")
            with open(probe, "w", encoding="utf-8") as f:
                f.write("ok")
            os.remove(probe)
            return candidate
        except Exception:
            continue
    raise PermissionError(
        f"Cannot write app data in either '{local_candidate}' or '{fallback_candidate}'."
    )

def get_app_data_dir():
    global _APP_DATA_DIR
    if _APP_DATA_DIR:
        return _APP_DATA_DIR
    _APP_DATA_DIR = _resolve_app_data_dir()
    return _APP_DATA_DIR

def get_config_file_path():
    return os.path.join(get_app_data_dir(), CONFIG_FILE_NAME)

def _ensure_app_dirs():
    os.makedirs(os.path.join(get_app_data_dir(), "working_copies"), exist_ok=True)

def get_master_db_path():
    _ensure_app_dirs()
    return os.path.join(get_app_data_dir(), "tool_leads.db")

def connect_master():
    path = get_master_db_path()
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    return conn

def init_master_schema(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            original_path TEXT NOT NULL,
            working_path TEXT NOT NULL UNIQUE,
            label TEXT,
            imported_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id INTEGER NOT NULL,
            email TEXT NOT NULL,
            full_name TEXT,
            number TEXT,
            sent INTEGER NOT NULL DEFAULT 0,
            replied INTEGER NOT NULL DEFAULT 0,
            replied_at TEXT,
            UNIQUE(import_id, email),
            FOREIGN KEY (import_id) REFERENCES imports(id)
        );
        CREATE INDEX IF NOT EXISTS idx_leads_email ON leads(email);

        -- Cache phone->country so we don't repeatedly call the external API.
        CREATE TABLE IF NOT EXISTS phone_country_cache (
            number_sanitized TEXT PRIMARY KEY,
            country_code TEXT NOT NULL,
            country_label TEXT NOT NULL,
            checked_at TEXT NOT NULL
        );
    """)
    conn.commit()

def ensure_contact_sent_column(conn):
    try:
        conn.execute("ALTER TABLE contacts ADD COLUMN sent INTEGER DEFAULT 0")
        conn.commit()
    except sqlite3.OperationalError:
        pass

def ensure_contact_replied_columns(conn):
    try:
        conn.execute("ALTER TABLE contacts ADD COLUMN replied INTEGER DEFAULT 0")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("ALTER TABLE contacts ADD COLUMN replied_at TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass

def ensure_contact_tracking_columns(conn):
    """Original lead .db files may lack sent/replied; add columns on the working copy only."""
    ensure_contact_sent_column(conn)
    ensure_contact_replied_columns(conn)

def normalize_email(addr):
    return (addr or "").strip().lower()

def register_working_database(original_path, working_path, label=None):
    """Register an existing SQLite working file in the master DB and sync leads (no file copy)."""
    _ensure_app_dirs()
    original_path = os.path.abspath(original_path)
    working_path = os.path.abspath(working_path)
    if not os.path.isfile(working_path):
        raise FileNotFoundError(f"Working database missing: {working_path}")
    conn_m = connect_master()
    init_master_schema(conn_m)
    lab = label or os.path.basename(original_path)
    ins = conn_m.execute(
        "INSERT INTO imports (original_path, working_path, label, imported_at) VALUES (?,?,?,?)",
        (original_path, working_path, lab, datetime.now().isoformat()),
    )
    import_id = ins.lastrowid
    _sync_working_file_to_master(conn_m, working_path, import_id)
    conn_m.commit()
    conn_m.close()
    return import_id, working_path

def import_user_database(original_path, label=None):
    """Copy the user's .db to the tool data dir and register it in the master DB."""
    _ensure_app_dirs()
    original_path = os.path.abspath(original_path)
    if not os.path.isfile(original_path):
        raise FileNotFoundError(f"Not a file: {original_path}")
    token = uuid.uuid4().hex[:12]
    base = os.path.splitext(os.path.basename(original_path))[0]
    working_path = os.path.join(get_app_data_dir(), "working_copies", f"{base}_{token}.db")
    shutil.copy2(original_path, working_path)
    return register_working_database(original_path, working_path, label)

def _excel_header_norm(cell):
    if cell is None:
        return ""
    return re.sub(r"\s+", " ", str(cell).strip().lower())

_EMAIL_RE = re.compile(r"^[A-Z0-9._%+\-']+@[A-Z0-9.\-]+\.[A-Z]{2,}$", re.IGNORECASE)

def _is_email_like(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    if not s or " " in s or "@" not in s:
        return False
    return _EMAIL_RE.match(s) is not None

def _digits_only(val) -> str:
    if val is None:
        return ""
    return re.sub(r"\D+", "", str(val))

def _is_phone_like(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    if "@" in s:
        return False
    d = _digits_only(s)
    return len(d) >= 7

def _excel_cell_to_int01(val, default=0):
    if val is None or val == "":
        return default
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, (int, float)):
        return 1 if int(val) != 0 else 0
    s = str(val).strip().lower()
    if s in ("1", "true", "yes", "y", "x"):
        return 1
    if s in ("0", "false", "no", "n"):
        return 0
    return default

def _score_excel_columns(header_vals, sample_rows, max_cols: int):
    """
    Identify email/phone/name columns even when headers don't match exactly.
    Uses a weighted score from header keywords and sample cell values.
    """
    headers = [_excel_header_norm(c) for c in (header_vals or ())]
    idx = {
        "full_name": None,
        "first_name": None,
        "last_name": None,
        "email": None,
        "number": None,
        "sent": None,
        "replied": None,
    }

    # Known status columns, when explicitly named
    for i, h in enumerate(headers[:max_cols]):
        if not h:
            continue
        if idx["sent"] is None and h == "sent":
            idx["sent"] = i
        if idx["replied"] is None and h in ("replied", "reply"):
            idx["replied"] = i
        if idx["first_name"] is None and (
            (("first" in h) and ("name" in h))
            or h in ("firstname", "given name", "givenname", "given")
        ):
            idx["first_name"] = i
        if idx["last_name"] is None and (
            (("last" in h) and ("name" in h))
            or h in ("lastname", "surname", "family name", "familyname", "last")
        ):
            idx["last_name"] = i
        if idx["full_name"] is None and (("full" in h) and ("name" in h)):
            idx["full_name"] = i

    email_scores = [0.0] * max_cols
    phone_scores = [0.0] * max_cols
    name_scores = [0.0] * max_cols

    for i in range(max_cols):
        h = headers[i] if i < len(headers) else ""
        if any(x in h for x in ("email", "e-mail", "e mail", "mail")):
            email_scores[i] += 3.0
        if any(x in h for x in ("phone", "mobile", "tel", "telephone", "cell", "whatsapp")) or h == "number":
            phone_scores[i] += 3.0
        if any(x in h for x in ("full name", "fullname", "contact name", "client name", "lead name")) or h in ("name", "contact"):
            name_scores[i] += 2.0
        if idx.get("first_name") == i:
            name_scores[i] += 1.5
        if idx.get("last_name") == i:
            name_scores[i] += 1.5

    for row in sample_rows or ():
        if not row:
            continue
        for i in range(max_cols):
            val = row[i] if i < len(row) else None
            if _is_email_like(val):
                email_scores[i] += 1.0
            if _is_phone_like(val):
                phone_scores[i] += 1.0
            if val is not None:
                s = str(val).strip()
                if s and not _is_email_like(s):
                    d = _digits_only(s)
                    if len(d) <= 2 and len(s) >= 3:
                        name_scores[i] += 0.2

    def best_index(scores):
        best_i, best_v = None, -1e9
        for i, v in enumerate(scores):
            if v > best_v:
                best_i, best_v = i, v
        return best_i, best_v

    email_i, email_v = best_index(email_scores)
    phone_i, phone_v = best_index(phone_scores)

    # Avoid mixing email + phone: ensure different columns
    if email_i is not None and phone_i is not None and email_i == phone_i:
        if email_v >= phone_v:
            phone_scores[email_i] = -1e9
            phone_i, phone_v = best_index(phone_scores)
        else:
            email_scores[phone_i] = -1e9
            email_i, email_v = best_index(email_scores)

    idx["email"] = email_i if (email_i is not None and email_v >= 2.0) else None
    idx["number"] = phone_i if (phone_i is not None and phone_v >= 2.0) else None

    forbidden = {idx["email"], idx["number"], idx["sent"], idx["replied"]}
    best_name_i, best_name_v = None, -1e9
    for i, v in enumerate(name_scores):
        if i in forbidden:
            continue
        if v > best_name_v:
            best_name_i, best_name_v = i, v
    if best_name_i is not None and best_name_v > 0.5:
        idx["full_name"] = best_name_i
    else:
        for i in range(max_cols):
            if i not in forbidden:
                idx["full_name"] = i
                break

    return idx

def _map_excel_columns(header_vals, sample_rows):
    """Map header row to column indices: full_name/first/last_name, email, number, sent, replied."""
    max_cols = max(len(header_vals or ()), max((len(r) for r in (sample_rows or ())), default=0), 0)
    max_cols = min(max_cols, 80)
    idx = _score_excel_columns(header_vals or (), sample_rows or (), max_cols)

    # Conservative fallback to A/B/C only if sample suggests column B is emails
    if idx["email"] is None and max_cols >= 3:
        col1_em = 0
        for r in (sample_rows or ())[:20]:
            if len(r) > 1 and _is_email_like(r[1]):
                col1_em += 1
        if col1_em >= 3:
            idx["full_name"], idx["email"], idx["number"] = 0, 1, 2
            if max_cols >= 5:
                idx["sent"], idx["replied"] = 3, 4
            elif max_cols >= 4:
                idx["sent"] = 3

    return idx

def write_excel_to_contacts_sqlite(excel_path, sqlite_path):
    """
    Scan all sheets to find the best match.
    Assumes row 1 is a header row, but header text can vary widely.
    Writes contacts table: full_name, number, email, sent, replied, replied_at.
    Returns number of rows inserted (rows with non-empty email).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "The openpyxl package is required for Excel import. Install with: pip install openpyxl"
        ) from e
    ext = os.path.splitext(excel_path)[1].lower()
    if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raise ValueError(
            "Supported Excel types: .xlsx / .xlsm. Save older .xls files as .xlsx in Excel."
        )
    if os.path.isfile(sqlite_path):
        os.remove(sqlite_path)
    conn = None
    wb = None
    try:
        conn = sqlite3.connect(sqlite_path)
        conn.execute(
            """
            CREATE TABLE contacts (
                full_name TEXT,
                number TEXT,
                email TEXT,
                sent INTEGER DEFAULT 0,
                replied INTEGER DEFAULT 0,
                replied_at TEXT
            )
            """
        )
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        best = None  # (score, ws, header_vals, sample_rows, col)

        for ws in wb.worksheets:
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_vals = next(rows_iter)
            except StopIteration:
                continue

            sample_rows = []
            for _ in range(50):
                try:
                    r = next(rows_iter)
                except StopIteration:
                    break
                sample_rows.append(r)

            col = _map_excel_columns(header_vals, sample_rows)
            if col.get("email") is None:
                continue

            email_hits = 0
            phone_hits = 0
            for r in sample_rows:
                if not r:
                    continue
                ix_email = col.get("email")
                if ix_email is not None and ix_email < len(r) and _is_email_like(r[ix_email]):
                    email_hits += 1
                ix_phone = col.get("number")
                if ix_phone is not None and ix_phone < len(r) and _is_phone_like(r[ix_phone]):
                    phone_hits += 1

            # Strongly prefer lots of email-like values; phone is a secondary tie-breaker
            score = (email_hits * 10.0) + (phone_hits * 1.0)
            if best is None or score > best[0]:
                best = (score, ws, header_vals, sample_rows, col)

        if best is None:
            raise ValueError(
                "Could not find an Email column in any sheet. "
                "Make sure row 1 contains headers (or at least that one sheet has email-like values)."
            )

        _, ws, header_vals, sample_rows, col = best
        inserted = 0
        batch = []
        def process_row(row):
            nonlocal inserted
            if row is None:
                return

            def gv(key):
                ix = col[key]
                if ix is None or ix >= len(row):
                    return None
                return row[ix]

            email_raw = gv("email")
            em = normalize_email(email_raw or "")
            if not em:
                return
            first = gv("first_name")
            last = gv("last_name")
            if (first is not None and str(first).strip() != "") or (last is not None and str(last).strip() != ""):
                first_s = str(first).strip() if first is not None else ""
                last_s = str(last).strip() if last is not None else ""
                full_name = (first_s + (" " if (first_s and last_s) else "") + last_s).strip()
            else:
                fn = gv("full_name")
                full_name = (str(fn).strip() if fn is not None else "") or ""
            num = gv("number")
            number = (str(num).strip() if num is not None else "") or ""
            sent = _excel_cell_to_int01(gv("sent"), 0)
            replied = _excel_cell_to_int01(gv("replied"), 0)
            rep_at = None
            batch.append((full_name, number, em, sent, replied, rep_at))
            inserted += 1

        # Process the sampled rows (already collected during scan)
        for row in sample_rows:
            process_row(row)

        # Then process remaining rows from the chosen worksheet
        rows_iter = ws.iter_rows(values_only=True)
        try:
            _ = next(rows_iter)  # header row
        except StopIteration:
            rows_iter = iter(())
        # Discard the same number of sampled rows so we don't double-import
        for _ in range(len(sample_rows)):
            try:
                _ = next(rows_iter)
            except StopIteration:
                break

        for row in rows_iter:
            process_row(row)
        if batch:
            conn.executemany(
                "INSERT INTO contacts (full_name, number, email, sent, replied, replied_at) VALUES (?,?,?,?,?,?)",
                batch,
            )
        conn.commit()
        return inserted
    except Exception:
        if conn:
            conn.close()
            conn = None
        try:
            if os.path.isfile(sqlite_path):
                os.remove(sqlite_path)
        except OSError:
            pass
        raise
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        if conn is not None:
            conn.close()

def import_excel_as_leads(original_path, label=None):
    """Convert Excel to a working .db in app data and register it (original Excel is never modified)."""
    _ensure_app_dirs()
    original_path = os.path.abspath(original_path)
    if not os.path.isfile(original_path):
        raise FileNotFoundError(f"Not a file: {original_path}")
    token = uuid.uuid4().hex[:12]
    base = os.path.splitext(os.path.basename(original_path))[0]
    working_path = os.path.join(get_app_data_dir(), "working_copies", f"{base}_excel_{token}.db")
    n = write_excel_to_contacts_sqlite(original_path, working_path)
    if n == 0:
        try:
            os.remove(working_path)
        except OSError:
            pass
        raise ValueError(
            "No rows with an email address were imported. "
            "Check the header row and that data starts on row 2."
        )
    return register_working_database(original_path, working_path, label)

def _sync_working_file_to_master(conn_m, working_path, import_id):
    """Upsert all contacts from a working clone into master.leads."""
    conn_w = sqlite3.connect(working_path)
    ensure_contact_tracking_columns(conn_w)
    cur = conn_w.cursor()
    cur.execute(
        """
        SELECT full_name, number, email,
               COALESCE(sent, 0) AS s,
               COALESCE(replied, 0) AS r,
               replied_at
        FROM contacts
        WHERE email IS NOT NULL AND TRIM(COALESCE(email, '')) != ''
        """
    )
    rows = cur.fetchall()
    conn_w.close()
    for full_name, number, email, sent, replied, rep_at in rows:
        e = normalize_email(email)
        if not e:
            continue
        conn_m.execute(
            """
            INSERT INTO leads (import_id, email, full_name, number, sent, replied, replied_at)
            VALUES (?,?,?,?,?,?, ?)
            ON CONFLICT(import_id, email) DO UPDATE SET
                full_name = excluded.full_name,
                number = excluded.number,
                sent = excluded.sent,
                replied = MAX(COALESCE(excluded.replied, 0), COALESCE(replied, 0)),
                replied_at = CASE
                    WHEN MAX(COALESCE(excluded.replied, 0), COALESCE(replied, 0)) != 0
                    THEN COALESCE(excluded.replied_at, replied_at)
                    ELSE NULL
                END
            """,
            (
                import_id,
                e,
                full_name,
                number,
                1 if sent else 0,
                1 if replied else 0,
                rep_at,
            ),
        )

def list_imports():
    conn = connect_master()
    init_master_schema(conn)
    rows = conn.execute(
        "SELECT id, original_path, working_path, label, imported_at FROM imports ORDER BY id DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_working_path_for_import(import_id):
    conn = connect_master()
    init_master_schema(conn)
    row = conn.execute("SELECT working_path FROM imports WHERE id=?", (import_id,)).fetchone()
    conn.close()
    return row[0] if row else None

def get_stats_for_import(import_id):
    """Counts come from the working .db `contacts` table. 'total' = rows with usable email (same as outreach)."""
    wp = get_working_path_for_import(import_id)
    if not wp or not os.path.isfile(wp):
        return {
            "total": 0,
            "sent": 0,
            "replied": 0,
            "left": 0,
            "total_rows": 0,
            "no_email": 0,
        }
    conn = sqlite3.connect(wp)
    ensure_contact_tracking_columns(conn)
    cur = conn.cursor()
    total_rows = cur.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
    base = "FROM contacts WHERE email IS NOT NULL AND TRIM(COALESCE(email, '')) != ''"
    total = cur.execute(f"SELECT COUNT(*) {base}").fetchone()[0]
    sent = cur.execute(
        f"SELECT COUNT(*) {base} AND COALESCE(sent, 0) != 0"
    ).fetchone()[0]
    replied = cur.execute(
        f"SELECT COUNT(*) {base} AND COALESCE(replied, 0) != 0"
    ).fetchone()[0]
    left = total - sent
    no_email = max(0, total_rows - total)
    conn.close()
    return {
        "total": total,
        "sent": sent,
        "replied": replied,
        "left": left,
        "total_rows": total_rows,
        "no_email": no_email,
    }

def get_import_id_for_working_path(working_path):
    working_path = os.path.abspath(working_path)
    conn = connect_master()
    init_master_schema(conn)
    row = conn.execute("SELECT id FROM imports WHERE working_path=?", (working_path,)).fetchone()
    conn.close()
    return row[0] if row else None

def add_lead_to_import(import_id, full_name, number, raw_email):
    email_n = normalize_email(raw_email)
    if not email_n:
        raise ValueError("Email is required.")
    conn = connect_master()
    init_master_schema(conn)
    row = conn.execute("SELECT working_path FROM imports WHERE id=?", (import_id,)).fetchone()
    if not row:
        conn.close()
        raise ValueError("Import not found.")
    working_path = row[0]
    conn.close()

    conn_w = sqlite3.connect(working_path)
    ensure_contact_tracking_columns(conn_w)
    conn_w.execute(
        "INSERT INTO contacts (full_name, number, email, sent, replied, replied_at) VALUES (?,?,?,0,0,NULL)",
        (full_name or "", number or "", email_n,),
    )
    conn_w.commit()
    conn_w.close()

    conn_m = connect_master()
    conn_m.execute(
        """
        INSERT INTO leads (import_id, email, full_name, number, sent, replied, replied_at)
        VALUES (?,?,?,?,0,0,NULL)
        ON CONFLICT(import_id, email) DO UPDATE SET
            full_name = excluded.full_name,
            number = excluded.number
        """,
        (import_id, email_n, full_name or "", number or ""),
    )
    conn_m.commit()
    conn_m.close()

def mark_replied_everywhere(email_addr):
    """Set replied on working .db row(s) and master leads (for sent matches)."""
    e = normalize_email(email_addr)
    if not e:
        return
    ts = datetime.now().isoformat()
    conn_m = connect_master()
    init_master_schema(conn_m)
    paths = conn_m.execute(
        """
        SELECT DISTINCT i.working_path FROM leads l
        JOIN imports i ON l.import_id = i.id
        WHERE l.email = ? AND l.sent = 1
        """,
        (e,),
    ).fetchall()
    for row in paths:
        wp = row[0]
        if not wp or not os.path.isfile(wp):
            continue
        cw = sqlite3.connect(wp)
        ensure_contact_tracking_columns(cw)
        cw.execute(
            "UPDATE contacts SET replied=1, replied_at=? WHERE lower(trim(email))=?",
            (ts, e),
        )
        cw.commit()
        cw.close()
    conn_m.execute(
        "UPDATE leads SET replied=1, replied_at=? WHERE email=? AND sent=1",
        (ts, e),
    )
    conn_m.commit()
    conn_m.close()

def remove_import(import_id):
    """Remove import record, master leads, and delete the working .db file."""
    conn = connect_master()
    init_master_schema(conn)
    row = conn.execute("SELECT working_path FROM imports WHERE id=?", (import_id,)).fetchone()
    if not row:
        conn.close()
        raise ValueError("Import not found.")
    wp = row[0]
    conn.execute("DELETE FROM leads WHERE import_id=?", (import_id,))
    conn.execute("DELETE FROM imports WHERE id=?", (import_id,))
    conn.commit()
    conn.close()
    if wp and os.path.isfile(wp):
        try:
            os.remove(wp)
        except OSError:
            pass

def resync_import_from_working(import_id):
    """Copy contacts from working .db into master.leads (webhook lookup)."""
    wp = get_working_path_for_import(import_id)
    if not wp or not os.path.isfile(wp):
        return
    conn_m = connect_master()
    init_master_schema(conn_m)
    _sync_working_file_to_master(conn_m, wp, import_id)
    conn_m.commit()
    conn_m.close()

def lookup_contact_master(email_addr):
    """Resolve lead for webhook: prefer a row we actually emailed."""
    e = normalize_email(email_addr)
    if not e:
        return None
    conn = connect_master()
    init_master_schema(conn)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT full_name, number, email FROM leads
        WHERE email=? AND sent=1
        ORDER BY id DESC LIMIT 1
        """,
        (e,),
    )
    row = cur.fetchone()
    if not row:
        cur.execute(
            """
            SELECT full_name, number, email FROM leads
            WHERE email=?
            ORDER BY id DESC LIMIT 1
            """,
            (e,),
        )
        row = cur.fetchone()
    conn.close()
    return tuple(row) if row else None

# ─────────────────────────────────────────────
# CONFIG HELPERS
# ─────────────────────────────────────────────
def load_config():
    # interpolation=None so email templates can contain "%" without errors
    c = configparser.ConfigParser(interpolation=None)
    config_path = get_config_file_path()
    if os.path.exists(config_path):
        c.read(config_path, encoding="utf-8")
    return c

def save_config(data: dict):
    c = load_config()
    for section, values in data.items():
        if not c.has_section(section):
            c.add_section(section)
        for k, v in values.items():
            c.set(section, k, str(v))
    config_path = get_config_file_path()
    with open(config_path, "w", encoding="utf-8") as f:
        c.write(f)

def cfg(section, key, fallback=""):
    c = load_config()
    if not c.has_section(section):
        return fallback
    return c.get(section, key, fallback=fallback)

# ─────────────────────────────────────────────
# EMAIL ENGINE
# ─────────────────────────────────────────────
def get_smtp():
    return smtplib.SMTP_SSL(cfg("SMTP","host","smtp.alexhost.com"), int(cfg("SMTP","port","465")))


def probe_rcpt_via_configured_smtp(
    email: str, mail_from: str, log_fn=None
) -> Optional[bool]:
    """
    If direct MX RCPT fails (common when outbound TCP/25 is blocked), run the same
    MAIL/RCPT sequence on the configured outbound SMTP (e.g. 465 SSL) after AUTH.
    """
    user = (cfg("SMTP", "user") or "").strip()
    password = cfg("SMTP", "password") or ""
    if not user or not password:
        _verification_log(
            "   ⚠️ SMTP credentials incomplete — skip relay RCPT fallback",
            log_fn,
        )
        return None
    try:
        with get_smtp() as s:
            s.login(user, password)
            return _smtp_handshake_rcpt(s, email, mail_from)
    except Exception as ex:
        _verification_log(f"   ⚠️ Relay RCPT fallback error: {ex}", log_fn)
        return None


def get_imap():
    m = imaplib.IMAP4_SSL(cfg("IMAP","host","imap.alexhost.com"), int(cfg("IMAP","port","993")))
    m.login(cfg("SMTP","user"), cfg("SMTP","password"))
    return m

def _imap_append_to_sent(raw_msg: bytes):
    """
    Best-effort: append a copy to a Sent folder so you can verify the message left.
    Not all servers expose the same folder names; we try common ones.
    """
    folders = ["Sent", "Sent Items", "INBOX.Sent", "INBOX.Sent Items"]
    imap = get_imap()
    try:
        # Discover available mailboxes (optional, but helps)
        try:
            _, data = imap.list()
            if data:
                names = []
                for line in data:
                    if not line:
                        continue
                    s = line.decode(errors="ignore")
                    # mailbox name is after last quote or last space; keep it simple
                    m = re.search(r'"([^"]+)"\s*$', s)
                    if m:
                        names.append(m.group(1))
                    else:
                        parts = s.split()
                        if parts:
                            names.append(parts[-1])
                # Prefer any mailbox that contains "sent"
                sent_like = [n for n in names if "sent" in (n or "").lower()]
                folders = sent_like + folders
        except Exception:
            pass

        for f in folders:
            if not f:
                continue
            try:
                imap.append(f, None, None, raw_msg)
                return True, f
            except Exception:
                continue
        return False, "no sent folder worked"
    finally:
        try:
            imap.logout()
        except Exception:
            pass

def send_email(to_email, subject, body, log_fn=None):
    user = cfg("SMTP","user")
    password = cfg("SMTP","password")
    msg = MIMEMultipart("alternative")
    msg["From"] = f"{cfg('SMTP','display_name',user)} <{user}>"
    msg["To"] = to_email
    msg["Subject"] = subject
    msg["Message-ID"] = email.utils.make_msgid()
    msg["Date"] = email.utils.formatdate(localtime=True)
    msg.attach(MIMEText(body, "plain"))

    ok_syntax, reason = verify_email(to_email, log_fn=log_fn, mail_from=user)
    if not ok_syntax:
        if log_fn:
            log_fn(f"❌ Failed {to_email}: {reason}")
        return False

    try:
        with get_smtp() as s:
            s.login(user, password)
            refused = s.sendmail(user, [to_email], msg.as_string())
        if refused:
            # refused is a dict: {recipient: (code, message)}
            try:
                code, msg_txt = list(refused.values())[0]
                reason = f"SMTP refused recipient ({code}): {msg_txt}"
            except Exception:
                reason = f"SMTP refused recipient: {refused}"
            if log_fn:
                log_fn(f"❌ Failed {to_email}: {reason}")
            return False
        if log_fn:
            log_fn(f"✅ Sent to {to_email} (accepted by SMTP) — msgid {msg['Message-ID']}")
        # Best-effort: save a copy to Sent (helps debug “accepted but not received”)
        try:
            ok_sent, sent_folder = _imap_append_to_sent(msg.as_bytes())
            if log_fn:
                if ok_sent:
                    log_fn(f"   📁 Saved a copy to IMAP folder: {sent_folder}")
                else:
                    log_fn(f"   ⚠️ Could not save to Sent: {sent_folder}")
        except Exception as ex:
            if log_fn:
                log_fn(f"   ⚠️ Could not save to Sent: {ex}")
        return True
    except Exception as e:
        if log_fn: log_fn(f"❌ Failed {to_email}: {e}")
        return False

def _sanitize_phone_country_key(raw_number) -> str:
    """Cache key: digits only (no '+' / dashes / spaces)."""
    return re.sub(r"\D+", "", str(raw_number or ""))

def _country_code_to_label(cc: str) -> str:
    cc = (cc or "").strip().upper()
    if cc == "GB":
        return "England"
    if cc == "CA":
        return "Canada"
    if cc == "US":
        return "United States"
    if cc == "AU":
        return "Australia"
    if cc == "FR":
        return "France"
    if cc == "DE":
        return "Germany"
    if cc == "ES":
        return "Spain"
    if cc == "IT":
        return "Italy"
    if cc == "NL":
        return "Netherlands"
    if cc == "BE":
        return "Belgium"
    return cc or "Unknown"

def _build_phone_candidates_for_api(raw_number: str, digits_key: str):
    """
    Build candidate phone strings for libphonenumberapi.com.
    We try explicit '+' first, then common country-prefix variants for national numbers.
    """
    s = (raw_number or "").strip()
    if not s and not digits_key:
        return []
    s2 = re.sub(r"[^\d+]", "", s)
    if s2.startswith("00"):
        s2 = "+" + s2[2:]
    candidates = []

    if s2.startswith("+") and s2 not in candidates:
        candidates.append(s2)
    if digits_key:
        candidates.append("+" + digits_key)
        if len(digits_key) == 10:
            candidates.append("+1" + digits_key)
            candidates.append("+44" + digits_key)
        elif len(digits_key) == 9:
            candidates.append("+44" + digits_key)

    # Deduplicate preserving order
    seen = set()
    out = []
    for c in candidates:
        c = c.strip()
        if not c or c in seen:
            continue
        seen.add(c)
        out.append(c)
    return out

def resolve_phone_location_label(conn_m, raw_number: str) -> str:
    """
    Resolve phone -> country label using the persistent cache.
    If not cached, call libphonenumberapi.com once and store the result.
    """
    key = _sanitize_phone_country_key(raw_number)
    if not key:
        return "Unknown"

    try:
        row = conn_m.execute(
            "SELECT country_label FROM phone_country_cache WHERE number_sanitized=?",
            (key,),
        ).fetchone()
        if row and row[0]:
            return row[0]
    except Exception:
        pass

    api_base = "https://libphonenumberapi.com/api/phone-numbers/"
    resolved_country = ""
    candidates = _build_phone_candidates_for_api(raw_number, key)

    # Try candidates until one returns a country code.
    for cand in candidates[:6]:
        try:
            url = api_base + requests.utils.quote(cand, safe="")
            r = requests.get(url, timeout=15)
            if r.status_code != 200:
                continue
            data = r.json()
            cc = (data.get("country") or "").strip()
            if cc:
                resolved_country = cc
                break
        except Exception:
            continue

    label = _country_code_to_label(resolved_country)
    # Cache even unknown to avoid repeated external calls for the same key.
    try:
        conn_m.execute(
            "INSERT OR REPLACE INTO phone_country_cache (number_sanitized, country_code, country_label, checked_at) VALUES (?,?,?,?)",
            (key, resolved_country or "", label, datetime.utcnow().isoformat()),
        )
        conn_m.commit()
    except Exception:
        try:
            conn_m.rollback()
        except Exception:
            pass
    return label

# ─────────────────────────────────────────────
# WARMUP ENGINE
# ─────────────────────────────────────────────
WARMUP_PHRASES = [
    ("Checking in", "Hey, just wanted to reach out and say hello. Hope you're having a great week!"),
    ("Quick hello", "Hi there! Just testing this email setup. Everything looks good on my end."),
    ("Touch base", "Hi! Reaching out to connect. Feel free to reply anytime."),
    ("Hello from my desk", "Just a friendly message to keep things warm. Have a wonderful day!"),
    ("Friendly ping", "Hi! Dropping a quick note. Hope all is well with you."),
]

def run_warmup(warmup_emails, count, delay_min, delay_max, log_fn, stop_event):
    log_fn(f"🔥 Starting warmup — sending {count} emails to warmup list...")
    sent = 0
    for i in range(count):
        if stop_event.is_set():
            log_fn("⛔ Warmup stopped.")
            break
        target = warmup_emails[i % len(warmup_emails)].strip()
        subj, body = random.choice(WARMUP_PHRASES)
        ok = send_email(target, subj, body, log_fn)
        if ok:
            sent += 1
        wait = random.randint(delay_min, delay_max)
        log_fn(f"   ⏳ Waiting {wait}s before next warmup email...")
        for _ in range(wait):
            if stop_event.is_set(): break
            time.sleep(1)
    log_fn(f"🔥 Warmup complete. Sent {sent}/{count} emails.")

# ─────────────────────────────────────────────
# OUTREACH ENGINE
# ─────────────────────────────────────────────
def load_db_contacts(db_path):
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    ensure_contact_tracking_columns(conn)
    cur.execute("SELECT full_name, number, email FROM contacts WHERE email IS NOT NULL AND trim(email) != ''")
    rows = cur.fetchall()
    conn.close()
    return rows  # [(full_name, number, email), ...]

def mark_sent(db_path, email_addr):
    """Update the working clone only (never the user's original file). Also sync master.leads."""
    db_path = os.path.abspath(db_path)
    em = normalize_email(email_addr)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    ensure_contact_tracking_columns(conn)
    cur.execute("UPDATE contacts SET sent=1 WHERE lower(trim(email))=?", (em,))
    conn.commit()
    conn.close()
    iid = get_import_id_for_working_path(db_path)
    if iid is None:
        return
    conn_m = connect_master()
    init_master_schema(conn_m)
    conn_m.execute(
        "UPDATE leads SET sent=1 WHERE import_id=? AND email=?",
        (iid, em),
    )
    conn_m.commit()
    conn_m.close()

def get_unsent(db_path):
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    ensure_contact_tracking_columns(conn)
    cur.execute(
        "SELECT full_name, number, email FROM contacts WHERE email IS NOT NULL AND trim(email) != '' "
        "AND (sent IS NULL OR sent=0)"
    )
    rows = cur.fetchall()
    conn.close()
    return rows

def personalize(template, name):
    first = name.split()[0] if name else "there"
    return template.replace("{name}", first).replace("{full_name}", name or "there")

def run_outreach(
    db_path,
    subject,
    body_template,
    daily_limit,
    delay_min,
    delay_max,
    log_fn,
    stop_event,
    status_fn=None,
):
    contacts = get_unsent(db_path)
    log_fn(f"📋 {len(contacts)} unsent contacts found.")
    conn_m = connect_master()
    init_master_schema(conn_m)
    sent_today = 0
    for name, number, email_addr in contacts:
        if stop_event.is_set():
            log_fn("⛔ Outreach stopped.")
            break
        if sent_today >= daily_limit:
            log_fn(f"📅 Daily limit of {daily_limit} reached. Will continue tomorrow.")
            break
        body = personalize(body_template, name or "")
        subj = personalize(subject, name or "")
        location = resolve_phone_location_label(conn_m, number)
        log_fn(f"📍 Location: {location} — Sending to {email_addr}")
        if status_fn:
            try:
                status_fn(0)
            except Exception:
                pass
        ok = send_email(email_addr, subj, body, log_fn=log_fn)
        if ok:
            mark_sent(db_path, email_addr)
            sent_today += 1
            log_fn(f"✅ Sent to {email_addr} ({location})")
        else:
            log_fn(f"❌ Failed to {email_addr} ({location})")
        wait = random.randint(delay_min, delay_max)
        log_fn(f"   ⏳ Next email in {wait}s... ({sent_today}/{daily_limit} today)")
        for remaining in range(wait, 0, -1):
            if stop_event.is_set():
                break
            if status_fn:
                try:
                    status_fn(remaining)
                except Exception:
                    pass
            time.sleep(1)
    log_fn(f"✅ Outreach session done. Sent {sent_today} emails.")
    try:
        conn_m.close()
    except Exception:
        pass

# ─────────────────────────────────────────────
# REPLY MONITOR
# ─────────────────────────────────────────────
def send_discord_alert(webhook_url, contact, reply_snippet, location: str):
    full_name, number, email_addr = contact
    payload = {
        "embeds": [{
            "title": "📬 Reply Received!",
            "color": 0x00C896,
            "fields": [
                {"name": "👤 Name", "value": full_name or "Unknown", "inline": True},
                {"name": "📧 Email", "value": email_addr, "inline": False},
                {"name": "📞 Phone", "value": number or "N/A", "inline": True},
                {"name": "🌍 Location", "value": location or "Unknown", "inline": False},
                {"name": "💬 Reply Preview", "value": reply_snippet[:300] or "—", "inline": False},
            ],
            "timestamp": datetime.utcnow().isoformat()
        }]
    }
    try:
        r = requests.post(webhook_url, json=payload, timeout=10)
        return r.status_code == 204
    except Exception as e:
        return False

def extract_sender_email(from_header):
    match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', from_header)
    return match.group(0).lower() if match else None

def run_reply_monitor(webhook_url, check_interval, log_fn, stop_event, seen_ids: set):
    _ensure_app_dirs()
    conn = connect_master()
    init_master_schema(conn)
    log_fn("👀 Reply monitor started (lookup: master tool_leads.db).")
    while not stop_event.is_set():
        try:
            imap = get_imap()
            imap.select("INBOX")
            _, data = imap.search(None, "ALL")
            msg_ids = data[0].split()
            for mid in msg_ids:
                if mid in seen_ids:
                    continue
                seen_ids.add(mid)
                _, msg_data = imap.fetch(mid, "(RFC822)")
                raw = msg_data[0][1]
                msg = email.message_from_bytes(raw)
                from_h = msg.get("From", "")
                sender = extract_sender_email(from_h)
                if not sender:
                    continue
                contact = lookup_contact_master(sender)
                if contact:
                    location = resolve_phone_location_label(conn, contact[1])
                    body_text = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                body_text = part.get_payload(decode=True).decode(errors="ignore")
                                break
                    else:
                        body_text = msg.get_payload(decode=True).decode(errors="ignore")
                    log_fn(f"📬 Reply from {sender} ({contact[0]}) [{location}] — alerting Discord!")
                    ok = send_discord_alert(webhook_url, contact, body_text, location)
                    if ok:
                        mark_replied_everywhere(sender)
                        log_fn(f"   ✅ Discord notified for {sender}")
                    else:
                        log_fn(f"   ❌ Discord webhook failed for {sender}")
            imap.logout()
        except Exception as e:
            log_fn(f"⚠️ Monitor error: {e}")
        for _ in range(check_interval):
            if stop_event.is_set(): break
            time.sleep(1)
    log_fn("👀 Reply monitor stopped.")
    try:
        conn.close()
    except Exception:
        pass

# ─────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("OutreachPro — Scam Recovery Mailer")
        self.geometry("960x720")
        self.resizable(True, True)
        self.configure(bg="#0D0D0D")
        self._setup_styles()

        self.stop_warmup = threading.Event()
        self.stop_outreach = threading.Event()
        self.stop_monitor = threading.Event()
        self.seen_ids = set()
        self._outreach_paths = []  # parallel to outreach combobox labels
        self._outreach_busy = False
        self._outreach_thread = None
        self._monitor_busy = False
        self._monitor_thread = None
        self._db_loc_refresh_busy = False
        self._db_loc_refresh_thread = None

        _ensure_app_dirs()
        mc = connect_master()
        init_master_schema(mc)
        mc.close()

        self._build_ui()
        self._load_saved_config()
        self._refresh_import_lists(select_working=cfg("OUTREACH", "working_path", ""))
        self._save_after_id = None
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._wire_autosave_prefs()

    def _setup_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TNotebook", background="#0D0D0D", borderwidth=0)
        style.configure("TNotebook.Tab", background="#1A1A1A", foreground="#888", padding=[16,8],
                        font=("Courier New", 10, "bold"))
        style.map("TNotebook.Tab", background=[("selected","#00C896")], foreground=[("selected","#000")])
        style.configure("TFrame", background="#0D0D0D")
        style.configure("TLabel", background="#0D0D0D", foreground="#CCC", font=("Courier New", 10))
        style.configure("TEntry", fieldbackground="#1A1A1A", foreground="#00C896",
                        insertcolor="#00C896", font=("Courier New", 10))
        style.configure("TButton", background="#00C896", foreground="#000",
                        font=("Courier New", 10, "bold"), padding=[10,6])
        style.map("TButton", background=[("active","#00A07A")])
        style.configure("Danger.TButton", background="#FF4444", foreground="#FFF",
                        font=("Courier New", 10, "bold"), padding=[10,6])
        style.map("Danger.TButton", background=[("active","#CC2222")])
        style.configure("TSpinbox", fieldbackground="#1A1A1A", foreground="#00C896",
                        font=("Courier New", 10))
        style.configure("TCheckbutton", background="#0D0D0D", foreground="#CCC",
                        font=("Courier New", 10))
        style.configure("TCombobox", fieldbackground="#1A1A1A", foreground="#00C896",
                        bordercolor="#333", arrowcolor="#00C896", font=("Courier New", 10))

    def _lbl(self, parent, text, row, col, colspan=1, anchor="w"):
        ttk.Label(parent, text=text).grid(row=row, column=col, columnspan=colspan,
                                          sticky=anchor, padx=8, pady=4)

    def _entry(self, parent, row, col, width=35, show=None, colspan=1):
        e = ttk.Entry(parent, width=width, show=show)
        e.grid(row=row, column=col, columnspan=colspan, sticky="ew", padx=8, pady=4)
        return e

    def _spin(self, parent, row, col, from_, to, default):
        s = ttk.Spinbox(parent, from_=from_, to=to, width=10)
        s.set(default)
        s.grid(row=row, column=col, sticky="w", padx=8, pady=4)
        return s

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self, bg="#0D0D0D")
        hdr.pack(fill="x", padx=20, pady=(16,4))
        hdr_status_fr = tk.Frame(hdr, bg="#0D0D0D")
        hdr_status_fr.pack(side="right", padx=(8, 0))
        self.hdr_monitor_status = tk.Label(
            hdr_status_fr,
            text="Reply monitor: Idle",
            bg="#0D0D0D",
            fg="#666666",
            font=("Courier New", 10, "bold"),
        )
        self.hdr_monitor_status.pack(anchor="e")
        self.hdr_outreach_status = tk.Label(
            hdr_status_fr,
            text="Outreach: Idle",
            bg="#0D0D0D",
            fg="#666666",
            font=("Courier New", 10, "bold"),
        )
        self.hdr_outreach_status.pack(anchor="e")
        tk.Label(hdr, text="OUTREACH", bg="#0D0D0D", fg="#00C896",
                 font=("Courier New", 22, "bold")).pack(side="left")
        tk.Label(hdr, text="PRO", bg="#0D0D0D", fg="#FFFFFF",
                 font=("Courier New", 22, "bold")).pack(side="left")
        tk.Label(hdr, text="  //  scam recovery mailer", bg="#0D0D0D", fg="#444",
                 font=("Courier New", 11)).pack(side="left", padx=8)

        sep = tk.Frame(self, bg="#00C896", height=1)
        sep.pack(fill="x", padx=20, pady=(0,8))

        self.status_bar = tk.Frame(self, bg="#141414", highlightthickness=1, highlightbackground="#2a2a2a")
        self.status_bar.pack(side="bottom", fill="x")
        self.status_outreach = tk.Label(
            self.status_bar,
            text="Outreach: Idle",
            bg="#141414",
            fg="#666666",
            font=("Courier New", 10),
            anchor="w",
        )
        self.status_outreach.pack(side="left", padx=(16, 8), pady=8)
        tk.Label(self.status_bar, text="·", bg="#141414", fg="#444444", font=("Courier New", 10)).pack(
            side="left", padx=4, pady=8
        )
        self.status_monitor = tk.Label(
            self.status_bar,
            text="Reply monitor: Idle",
            bg="#141414",
            fg="#666666",
            font=("Courier New", 10),
            anchor="w",
        )
        self.status_monitor.pack(side="left", padx=(8, 16), pady=8)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=20, pady=(0, 8))
        self.notebook.bind("<<NotebookTabChanged>>", lambda _e: self._on_notebook_tab_changed())

        self._build_settings_tab(self.notebook)
        self._build_database_tab(self.notebook)
        self._build_warmup_tab(self.notebook)
        self._build_outreach_tab(self.notebook)
        self._build_monitor_tab(self.notebook)
        self._build_log_tab(self.notebook)

    def _on_notebook_tab_changed(self):
        # Persist settings after edits.
        self._schedule_prefs_save()

    # ── SETTINGS ──────────────────────────────
    def _build_settings_tab(self, nb):
        f = ttk.Frame(nb)
        nb.add(f, text="⚙  SETTINGS")
        f.columnconfigure(1, weight=1)

        self._lbl(f,"─── SMTP (Outgoing) ───",0,0,2)
        self._lbl(f,"SMTP Host:",1,0)
        self.smtp_host = self._entry(f,1,1); self.smtp_host.insert(0,"smtp.alexhost.com")
        self._lbl(f,"SMTP Port:",2,0)
        self.smtp_port = self._entry(f,2,1); self.smtp_port.insert(0,"465")
        self._lbl(f,"Email Address:",3,0)
        self.smtp_user = self._entry(f,3,1)
        self._lbl(f,"Password:",4,0)
        self.smtp_pass = self._entry(f,4,1,show="●")
        self._lbl(f,"Display Name:",5,0)
        self.smtp_name = self._entry(f,5,1)

        self._lbl(f,"─── IMAP (Incoming / Reply Detection) ───",6,0,2)
        self._lbl(f,"IMAP Host:",7,0)
        self.imap_host = self._entry(f,7,1); self.imap_host.insert(0,"imap.alexhost.com")
        self._lbl(f,"IMAP Port:",8,0)
        self.imap_port = self._entry(f,8,1); self.imap_port.insert(0,"993")

        self._lbl(f,"─── Discord ───",9,0,2)
        self._lbl(f,"Webhook URL:",10,0)
        self.discord_url = self._entry(f,10,1,width=60)

        ttk.Button(f, text="💾  Save all preferences", command=self._save_settings).grid(
            row=11, column=0, columnspan=2, pady=16)
        self._lbl(
            f,
            "Preferences auto-save when you switch tabs, after edits (short delay), and when closing.",
            12,
            0,
            2,
        )

    # ── DATABASE ──────────────────────────────
    def _build_database_tab(self, nb):
        f = ttk.Frame(nb)
        nb.add(f, text="🗄  DATABASE")
        f.columnconfigure(0, weight=1)
        f.columnconfigure(1, weight=1)

        top = ttk.Frame(f)
        top.grid(row=0, column=0, columnspan=2, sticky="ew", padx=8, pady=8)
        ttk.Button(top, text="➕  Import .db (saves a working copy)", command=self._import_database_dialog).pack(
            side="left", padx=4
        )
        ttk.Button(top, text="📊  Import Excel → .db", command=self._import_excel_dialog).pack(
            side="left", padx=4
        )
        ttk.Button(top, text="🔄  Refresh lists", command=lambda: self._refresh_import_lists()).pack(side="left", padx=4)
        ttk.Button(top, text="🗑  Remove selected import", style="Danger.TButton", command=self._remove_selected_import).pack(
            side="left", padx=4
        )

        self._lbl(f, "Imported lists (original file is never modified):", 1, 0, 2)
        lb_frame = tk.Frame(f, bg="#0D0D0D")
        lb_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=8, pady=4)
        f.rowconfigure(2, weight=1)
        self.db_imports_lb = tk.Listbox(
            lb_frame,
            height=8,
            bg="#1A1A1A",
            fg="#00C896",
            selectbackground="#00A07A",
            selectforeground="#000",
            font=("Courier New", 10),
            relief="flat",
            highlightthickness=0,
        )
        self.db_imports_lb.pack(fill="both", expand=True)
        self.db_imports_lb.bind("<<ListboxSelect>>", lambda _e: self._on_database_list_select())

        stat_fr = ttk.Frame(f)
        stat_fr.grid(row=3, column=0, columnspan=2, sticky="ew", padx=8, pady=8)
        self.db_stat_total = ttk.Label(stat_fr, text="Leads: —")
        self.db_stat_total.pack(side="left", padx=12)
        self.db_stat_sent = ttk.Label(stat_fr, text="Sent: —")
        self.db_stat_sent.pack(side="left", padx=12)
        self.db_stat_replied = ttk.Label(stat_fr, text="Replied: —")
        self.db_stat_replied.pack(side="left", padx=12)
        self.db_stat_left = ttk.Label(stat_fr, text="Left (not sent): —")
        self.db_stat_left.pack(side="left", padx=12)

        stat_note = ttk.Frame(f)
        stat_note.grid(row=4, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 4))
        self.db_stat_note = ttk.Label(
            stat_note,
            text="",
            font=("Courier New", 9),
            foreground="#888",
        )
        self.db_stat_note.pack(side="left")

        self._lbl(f, "Leads by location (from phone number country code):", 5, 0, 2)

        # Compact inline summary only (no table/tree view).
        self.db_loc_summary = ttk.Label(
            f,
            text="",
            font=("Courier New", 9),
            foreground="#888",
            wraplength=740,
        )
        self.db_loc_summary.grid(row=6, column=0, columnspan=2, sticky="ew", padx=8, pady=(4, 0))

        self._lbl(
            f,
            "ℹ  .db imports are copied into the tool data folder; Excel is converted to a new\n"
            "   .db (row 1 = headers: Full name, Email, Number, Sent, Replied). Supported:\n"
            "   .xlsx / .xlsm. Outreach updates the working copy; reply monitor uses master DB.\n"
            "   'Leads' = `contacts` rows with a non-empty email.",
            10,
            0,
            2,
        )

        self._db_list_import_ids = []

    # ── WARMUP ────────────────────────────────
    def _build_warmup_tab(self, nb):
        f = ttk.Frame(nb)
        nb.add(f, text="🔥  WARMUP")
        f.columnconfigure(1, weight=1)

        self._lbl(f,"Warmup Email List (one per line):",0,0,2)
        self.warmup_emails = tk.Text(f, height=6, bg="#1A1A1A", fg="#00C896",
                                     insertbackground="#00C896",
                                     font=("Courier New",10), relief="flat")
        self.warmup_emails.grid(row=1,column=0,columnspan=2,sticky="ew",padx=8,pady=4)

        self._lbl(f,"Emails to Send:",2,0)
        self.warmup_count = self._spin(f,2,1,1,500,30)
        self._lbl(f,"Min Delay (sec):",3,0)
        self.warmup_dmin = self._spin(f,3,1,10,600,60)
        self._lbl(f,"Max Delay (sec):",4,0)
        self.warmup_dmax = self._spin(f,4,1,10,600,180)

        bf = ttk.Frame(f)
        bf.grid(row=5,column=0,columnspan=2,pady=10)
        ttk.Button(bf, text="▶  Start Warmup", command=self._start_warmup).pack(side="left",padx=4)
        ttk.Button(bf, text="⏹  Stop", style="Danger.TButton", command=self._stop_warmup).pack(side="left",padx=4)

        self._lbl(f,"ℹ  Warmup sends natural-looking emails to trusted addresses to build\n"
                    "   your sender reputation before doing real outreach.",6,0,2)

    # ── OUTREACH ──────────────────────────────
    def _build_outreach_tab(self, nb):
        f = ttk.Frame(nb)
        nb.add(f, text="📧  OUTREACH")
        f.columnconfigure(1, weight=1)

        self._lbl(f, "Outreach list (working copy):", 0, 0)
        row0 = ttk.Frame(f)
        row0.grid(row=0, column=1, sticky="ew", padx=8, pady=4)
        self.outreach_import = ttk.Combobox(row0, width=52, state="readonly")
        self.outreach_import.pack(side="left")
        self.outreach_import.bind("<<ComboboxSelected>>", lambda _e: self._schedule_prefs_save())
        ttk.Button(row0, text="Import .db…", command=self._import_database_dialog).pack(side="left", padx=4)
        ttk.Button(row0, text="Import Excel…", command=self._import_excel_dialog).pack(side="left", padx=4)

        self._lbl(f,"Subject Line:",1,0)
        self.email_subject = self._entry(f,1,1,width=60)
        self.email_subject.insert(0,"Regarding your case — we may be able to help")

        self._lbl(f,"Email Body (use {name} for first name):",2,0,2)
        self.email_body = tk.Text(f, height=10, bg="#1A1A1A", fg="#00C896",
                                   insertbackground="#00C896",
                                   font=("Courier New",10), relief="flat", wrap="word")
        self.email_body.grid(row=3,column=0,columnspan=2,sticky="ew",padx=8,pady=4)
        self.email_body.insert("1.0",
            "Hi {name},\n\n"
            "My name is [YOUR NAME] and I work with a team that specialises in helping victims of financial fraud recover their lost funds.\n\n"
            "We understand how distressing it can be to lose money to a scam, and we want you to know that recovery may be possible.\n\n"
            "If you'd like to learn more about your options, simply reply to this email and one of our specialists will reach out to you personally.\n\n"
            "There is no obligation, and your enquiry is completely confidential.\n\n"
            "Kind regards,\n[YOUR NAME]"
        )

        self._lbl(f,"Daily Send Limit:",4,0)
        self.daily_limit = self._spin(f,4,1,1,1000,100)
        self._lbl(f,"Min Delay Between Emails (sec):",5,0)
        self.out_dmin = self._spin(f,5,1,30,3600,120)
        self._lbl(f,"Max Delay Between Emails (sec):",6,0)
        self.out_dmax = self._spin(f,6,1,30,3600,300)

        bf = ttk.Frame(f)
        bf.grid(row=7,column=0,columnspan=2,pady=10)
        ttk.Button(bf, text="▶  Start Outreach", command=self._start_outreach).pack(side="left",padx=4)
        ttk.Button(bf, text="⏹  Stop", style="Danger.TButton", command=self._stop_outreach).pack(side="left",padx=4)

    # ── MONITOR ───────────────────────────────
    def _build_monitor_tab(self, nb):
        f = ttk.Frame(nb)
        nb.add(f, text="👀  REPLY MONITOR")
        f.columnconfigure(1, weight=1)

        self._lbl(f, "Lead lookup:", 0, 0)
        mp = get_master_db_path()
        self.master_db_label = ttk.Label(f, text=mp, font=("Courier New", 9))
        self.master_db_label.grid(row=0, column=1, sticky="w", padx=8, pady=4)

        self._lbl(f,"Check Inbox Every (sec):",1,0)
        self.check_interval = self._spin(f,1,1,30,3600,120)

        bf = ttk.Frame(f)
        bf.grid(row=2,column=0,columnspan=2,pady=10)
        ttk.Button(bf, text="▶  Start Monitor", command=self._start_monitor).pack(side="left",padx=4)
        ttk.Button(bf, text="⏹  Stop", style="Danger.TButton", command=self._stop_monitor).pack(side="left",padx=4)

        self._lbl(
            f,
            "ℹ  The monitor uses the tool's master database (all imported leads). When a\n"
            "   reply matches an email there, it posts to Discord and marks the lead as\n"
            "   replied in the master database.",
            3,
            0,
            2,
        )

    # ── LEADS ─────────────────────────────────
    def _build_leads_tab(self, nb):
        # Leads tab removed; location breakdown is now shown in the Database tab.
        return

    def _phone_location_label(self, raw_number: str, default_region_hint: Optional[str] = None) -> str:
        """
        Best-effort location label from phone number.

        If the number doesn't include an explicit country prefix (no '+' / '00'), we parse it
        using `default_region_hint` to avoid mis-interpreting country codes.
        """
        s = (raw_number or "").strip()
        if not s:
            return "Unknown"

        # Keep digits and leading '+' if present
        s2 = re.sub(r"[^\d+]", "", s)
        if not s2:
            return "Unknown"

        explicit = s2.startswith("+") or s2.startswith("00")
        if s2.startswith("00"):
            # Convert 00<countrycode> -> +<countrycode>
            s2 = "+" + s2[2:]

        digits = re.sub(r"\D+", "", s2)  # after normalization, remove any '+'
        if not digits:
            return "Unknown"

        try:
            if explicit:
                num = phonenumbers.parse(s2, None)
                if not phonenumbers.is_possible_number(num):
                    return "Unknown"
            else:
                # Parse as a national number using the hint, if we have one.
                if default_region_hint:
                    num = phonenumbers.parse(digits, default_region_hint)
                    if not phonenumbers.is_possible_number(num):
                        # Fallback: try interpreting it as E.164 (may be wrong, but better than Unknown)
                        num = phonenumbers.parse("+" + digits, None)
                else:
                    num = phonenumbers.parse(digits, None)
                    if not phonenumbers.is_possible_number(num):
                        num = phonenumbers.parse("+" + digits, None)

                if not phonenumbers.is_possible_number(num):
                    return "Unknown"

            region = phonenumbers.region_code_for_number(num) or "Unknown"
        except Exception:
            return "Unknown"

        if region == "GB":
            return "England"
        if region == "FR":
            return "France"
        if region == "US":
            return "United States"
        if region == "CA":
            return "Canada"
        if region == "AU":
            return "Australia"
        if region == "DE":
            return "Germany"
        if region == "ES":
            return "Spain"
        if region == "IT":
            return "Italy"
        if region == "NL":
            return "Netherlands"
        if region == "BE":
            return "Belgium"
        return region

    def _clear_db_location_breakdown(self):
        try:
            self.db_loc_summary.config(text="")
        except Exception:
            pass

    def _refresh_db_location_breakdown(self, import_id):
        """Show location summary for the selected import (working .db).

        Uses libphonenumberapi.com to resolve phone -> country, and caches results
        in the master DB so the API is only called once per phone number.
        """
        if self._db_loc_refresh_busy:
            return

        self._db_loc_refresh_busy = True
        try:
            self.db_loc_summary.config(text="Loading location counts...")
        except Exception:
            pass

        def worker():
            counts = {}
            try:
                wp = get_working_path_for_import(import_id)
                if not wp or not os.path.isfile(wp):
                    raise FileNotFoundError("Working DB missing")

                # Load all phone numbers from the selected working DB.
                conn_w = sqlite3.connect(wp)
                ensure_contact_tracking_columns(conn_w)
                rows = conn_w.execute(
                    "SELECT number FROM contacts WHERE number IS NOT NULL AND TRIM(COALESCE(number,'')) != ''"
                ).fetchall()
                conn_w.close()

                # Prepare deduped list of phones for API caching.
                def sanitize_key(v):
                    return re.sub(r"\D+", "", str(v or ""))

                key_to_sample_raw = {}
                keys_in_file = set()
                key_freq = {}
                for r in rows:
                    raw = r[0]
                    key = sanitize_key(raw)
                    if not key:
                        continue
                    keys_in_file.add(key)
                    if key not in key_to_sample_raw:
                        key_to_sample_raw[key] = raw
                    key_freq[key] = key_freq.get(key, 0) + 1

                # Determine a lightweight default region hint (only for constructing candidate API inputs).
                default_region_hint = None
                try:
                    hint_counts = {}
                    candidate_regions = ["CA", "GB", "US", "AU", "FR", "DE", "ES", "IT", "NL", "BE"]
                    # Use a small sample for speed.
                    for key in list(keys_in_file)[:500]:
                        try:
                            digits = key
                            for cand in candidate_regions:
                                num = phonenumbers.parse(digits, cand)
                                if phonenumbers.is_possible_number(num):
                                    region = phonenumbers.region_code_for_number(num) or None
                                    if region:
                                        hint_counts[region] = hint_counts.get(region, 0) + 1
                        except Exception:
                            continue
                    if hint_counts:
                        default_region_hint = max(hint_counts.items(), key=lambda kv: kv[1])[0]
                except Exception:
                    default_region_hint = None

                # Open master DB and read cache for the phones we need.
                conn_m = connect_master()
                init_master_schema(conn_m)

                def label_from_country_code(cc: str) -> str:
                    cc = (cc or "").strip().upper()
                    if cc == "GB":
                        return "England"
                    if cc == "CA":
                        return "Canada"
                    if cc == "US":
                        return "United States"
                    if cc == "AU":
                        return "Australia"
                    if cc == "FR":
                        return "France"
                    if cc == "DE":
                        return "Germany"
                    if cc == "ES":
                        return "Spain"
                    if cc == "IT":
                        return "Italy"
                    if cc == "NL":
                        return "Netherlands"
                    if cc == "BE":
                        return "Belgium"
                    return cc or "Unknown"

                cache_labels = {}  # key -> label

                keys_list = list(keys_in_file)
                chunk_size = 500
                for i in range(0, len(keys_list), chunk_size):
                    chunk = keys_list[i:i + chunk_size]
                    if not chunk:
                        continue
                    q_marks = ",".join(["?"] * len(chunk))
                    rows_cache = conn_m.execute(
                        f"SELECT number_sanitized, country_label FROM phone_country_cache WHERE number_sanitized IN ({q_marks})",
                        tuple(chunk),
                    ).fetchall()
                    for k, lab in rows_cache:
                        cache_labels[k] = lab

                # Call API for any cache misses.
                missing_keys = [k for k in keys_in_file if k not in cache_labels]

                api_base = "https://libphonenumberapi.com/api/phone-numbers/"

                def build_api_candidates(sample_raw, digits_key):
                    # Return a list of phone strings to try (include '+').
                    s = str(sample_raw or "").strip()
                    s2 = re.sub(r"[^\d+]", "", s)
                    if s2.startswith("00"):
                        s2 = "+" + s2[2:]
                    candidates = []

                    if s2.startswith("+"):
                        candidates.append(s2)
                    # If digits already include country code (e.g. +1..., +44..., or US/CA '1' prefix in national form)
                    if len(digits_key) >= 11 and digits_key.startswith("1"):
                        candidates.append("+" + digits_key)

                    if default_region_hint in ("CA", "GB"):
                        # If we have a national number without country prefix, prepend the expected calling code.
                        if default_region_hint == "CA" and (len(digits_key) == 10 or not digits_key.startswith("1")):
                            candidates.append("+1" + digits_key[-10:])
                        if default_region_hint == "GB" and (len(digits_key) == 10 or not digits_key.startswith("44")):
                            candidates.append("+44" + digits_key[-10:])

                    # If the number looks like a national-format number (no +), try common candidates.
                    # This improves CA/GB disambiguation when we can't infer a country prefix.
                    if len(digits_key) == 10:
                        candidates.append("+1" + digits_key)
                        candidates.append("+44" + digits_key)
                    elif len(digits_key) == 9:
                        # Some UK formats are 9 digits after stripping leading 0.
                        candidates.append("+44" + digits_key)

                    # Final fallback: try interpreting as international by direct '+digits'.
                    candidates.append("+" + digits_key)

                    # Deduplicate while preserving order.
                    seen = set()
                    out = []
                    for c in candidates:
                        c = c.strip()
                        if not c or c in seen:
                            continue
                        seen.add(c)
                        out.append(c)
                    return out

                headers = {"User-Agent": "OutreachPro/1.0"}
                # Real-time progress + counts:
                # - Start from cached labels
                # - Count unresolved phones as "Pending"
                counts = {}
                pending_label = "Pending"
                pending_total = 0
                for k, freq in key_freq.items():
                    lab = cache_labels.get(k)
                    if lab:
                        counts[lab] = counts.get(lab, 0) + freq
                    else:
                        pending_total += freq
                if pending_total:
                    counts[pending_label] = pending_total

                total_missing = len(missing_keys)
                done_missing = 0
                counts_lock = Lock()
                cache_inserts = []  # (number_sanitized, country_code, country_label, checked_at)

                def set_summary_text(extra_prefix: str = ""):
                    try:
                        items = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0].lower()))
                        top_items = items[:12]
                        rest = len(items) - len(top_items)
                        summary = ", ".join([f"{loc} ({cnt})" for loc, cnt in top_items])
                        if rest > 0:
                            summary = summary + f", +{rest} more"
                        if extra_prefix:
                            summary = f"{extra_prefix} {summary}".strip()
                        return summary
                    except Exception:
                        return ""

                def resolve_one(key: str):
                    sample_raw = key_to_sample_raw.get(key, "")
                    digits_key = key
                    candidates = build_api_candidates(sample_raw, digits_key)
                    resolved_country = None
                    for cand in candidates:
                        try:
                            url = api_base + requests.utils.quote(cand, safe="")
                            r = requests.get(url, headers=headers, timeout=15)
                            if r.status_code != 200:
                                continue
                            data = r.json()
                            cc = (data.get("country") or "").strip()
                            if cc:
                                resolved_country = cc
                                break
                        except Exception:
                            continue
                    country_label = label_from_country_code(resolved_country) if resolved_country else "Unknown"
                    return key, resolved_country or "", country_label

                # Kick off a UI ticker that updates while we work.
                def ui_tick():
                    if not self._db_loc_refresh_busy:
                        return
                    with counts_lock:
                        prefix = f"Resolving {done_missing}/{total_missing}..."
                        live = set_summary_text(extra_prefix=prefix)
                    try:
                        self.db_loc_summary.config(text=live)
                    except Exception:
                        pass
                    self.after(350, ui_tick)

                self.after(0, ui_tick)

                if missing_keys:
                    with ThreadPoolExecutor(max_workers=10) as ex:
                        futures = [ex.submit(resolve_one, k) for k in missing_keys]
                        for fut in as_completed(futures):
                            try:
                                key, cc, lab = fut.result()
                            except Exception:
                                continue
                            freq = key_freq.get(key, 0)
                            with counts_lock:
                                done_missing += 1
                                # Move this key's rows from Pending -> resolved label
                                if pending_total and counts.get(pending_label, 0) > 0:
                                    counts[pending_label] = max(0, counts.get(pending_label, 0) - freq)
                                    if counts[pending_label] == 0:
                                        counts.pop(pending_label, None)
                                counts[lab] = counts.get(lab, 0) + freq
                            cache_labels[key] = lab
                            cache_inserts.append((key, cc, lab, datetime.utcnow().isoformat()))

                            # Flush cache inserts in batches so SQLite writes stay safe/fast.
                            if len(cache_inserts) >= 200:
                                try:
                                    conn_m.executemany(
                                        "INSERT OR REPLACE INTO phone_country_cache (number_sanitized, country_code, country_label, checked_at) VALUES (?,?,?,?)",
                                        cache_inserts,
                                    )
                                    conn_m.commit()
                                    cache_inserts.clear()
                                except Exception:
                                    conn_m.rollback()

                # Final cache flush
                if cache_inserts:
                    try:
                        conn_m.executemany(
                            "INSERT OR REPLACE INTO phone_country_cache (number_sanitized, country_code, country_label, checked_at) VALUES (?,?,?,?)",
                            cache_inserts,
                        )
                        conn_m.commit()
                    except Exception:
                        conn_m.rollback()

                # Final summary (no prefix)
                with counts_lock:
                    summary = set_summary_text()

                try:
                    conn_m.close()
                except Exception:
                    pass

            except Exception:
                summary = ""

            def apply_ui():
                try:
                    self.db_loc_summary.config(text=summary)
                except Exception:
                    pass
                self._db_loc_refresh_busy = False

            self.after(0, apply_ui)

        self._db_loc_refresh_thread = threading.Thread(target=worker, daemon=True)
        self._db_loc_refresh_thread.start()

    # ── LOG ───────────────────────────────────
    def _build_log_tab(self, nb):
        f = ttk.Frame(nb)
        nb.add(f, text="📋  LOG")
        self.log_box = scrolledtext.ScrolledText(
            f, bg="#0D0D0D", fg="#00C896", insertbackground="#00C896",
            font=("Courier New", 9), relief="flat", state="disabled"
        )
        self.log_box.pack(fill="both", expand=True, padx=8, pady=8)
        ttk.Button(f, text="🗑  Clear Log", command=self._clear_log).pack(pady=4)

    # ── ACTIONS ───────────────────────────────
    def log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        def _do():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", f"[{ts}] {msg}\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _do)

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0","end")
        self.log_box.configure(state="disabled")

    def _save_all_ui_prefs(self):
        """Write every tab’s fields to config.ini (same file as SMTP / Discord)."""
        save_config({
            "SMTP": {
                "host": self.smtp_host.get(),
                "port": self.smtp_port.get(),
                "user": self.smtp_user.get(),
                "password": self.smtp_pass.get(),
                "display_name": self.smtp_name.get(),
            },
            "IMAP": {
                "host": self.imap_host.get(),
                "port": self.imap_port.get(),
            },
            "DISCORD": {
                "webhook": self.discord_url.get(),
            },
            "WARMUP": {
                "emails": self.warmup_emails.get("1.0", "end").rstrip("\n"),
                "count": self.warmup_count.get(),
                "delay_min": self.warmup_dmin.get(),
                "delay_max": self.warmup_dmax.get(),
            },
            "OUTREACH": {
                "working_path": self._get_outreach_working_path(),
                "subject": self.email_subject.get(),
                "body": self.email_body.get("1.0", "end").rstrip("\n"),
                "daily_limit": self.daily_limit.get(),
                "delay_min": self.out_dmin.get(),
                "delay_max": self.out_dmax.get(),
            },
            "MONITOR": {
                "check_interval": self.check_interval.get(),
            },
        })

    def _schedule_prefs_save(self, _event=None):
        if self._save_after_id is not None:
            self.after_cancel(self._save_after_id)
        self._save_after_id = self.after(900, self._flush_prefs_save)

    def _flush_prefs_save(self):
        self._save_after_id = None
        try:
            self._save_all_ui_prefs()
        except Exception:
            pass

    def _wire_autosave_prefs(self):
        deb = self._schedule_prefs_save
        flush = self._flush_prefs_save
        for w in (
            self.smtp_host,
            self.smtp_port,
            self.smtp_user,
            self.smtp_pass,
            self.smtp_name,
            self.imap_host,
            self.imap_port,
            self.discord_url,
            self.email_subject,
        ):
            w.bind("<KeyRelease>", deb)
            w.bind("<FocusOut>", lambda _e: flush())
        self.warmup_emails.bind("<KeyRelease>", deb)
        self.warmup_emails.bind("<FocusOut>", lambda _e: flush())
        self.email_body.bind("<KeyRelease>", deb)
        self.email_body.bind("<FocusOut>", lambda _e: flush())
        for w in (
            self.warmup_count,
            self.warmup_dmin,
            self.warmup_dmax,
            self.daily_limit,
            self.out_dmin,
            self.out_dmax,
            self.check_interval,
        ):
            w.bind("<ButtonRelease-1>", deb)
            w.bind("<KeyRelease>", deb)
            w.bind("<FocusOut>", lambda _e: flush())

    def _on_close(self):
        try:
            if self._save_after_id is not None:
                self.after_cancel(self._save_after_id)
                self._save_after_id = None
            self._save_all_ui_prefs()
        except Exception:
            pass
        self.destroy()

    def _save_settings(self):
        try:
            self._save_all_ui_prefs()
        except Exception as ex:
            messagebox.showerror("Save failed", str(ex))
            return
        messagebox.showinfo("Saved", "All preferences were saved.")

    def _load_saved_config(self):
        c = load_config()

        def _set(widget, val):
            if val is not None and val != "":
                widget.delete(0, "end")
                widget.insert(0, val)

        # Backwards-compatible: if an old config still has Hostinger defaults,
        # switch it to AlexHost defaults.
        hostinger_default_smtp = "smtp.hostinger.com"
        alexhost_default_smtp = "smtp.alexhost.com"
        hostinger_default_imap = "imap.hostinger.com"
        alexhost_default_imap = "imap.alexhost.com"

        smtp_loaded = cfg("SMTP", "host", hostinger_default_smtp)
        if (smtp_loaded or "").strip().lower() == hostinger_default_smtp:
            smtp_loaded = alexhost_default_smtp
        _set(self.smtp_host, smtp_loaded)
        _set(self.smtp_port, cfg("SMTP", "port", "465"))
        _set(self.smtp_user, cfg("SMTP", "user"))
        _set(self.smtp_pass, cfg("SMTP", "password"))
        _set(self.smtp_name, cfg("SMTP", "display_name"))
        imap_loaded = cfg("IMAP", "host", hostinger_default_imap)
        if (imap_loaded or "").strip().lower() == hostinger_default_imap:
            imap_loaded = alexhost_default_imap
        _set(self.imap_host, imap_loaded)
        _set(self.imap_port, cfg("IMAP", "port", "993"))
        _set(self.discord_url, cfg("DISCORD", "webhook"))

        if c.has_section("WARMUP"):
            if c.has_option("WARMUP", "emails"):
                self.warmup_emails.delete("1.0", tk.END)
                self.warmup_emails.insert("1.0", c.get("WARMUP", "emails"))
            if c.has_option("WARMUP", "count"):
                self.warmup_count.set(c.get("WARMUP", "count"))
            if c.has_option("WARMUP", "delay_min"):
                self.warmup_dmin.set(c.get("WARMUP", "delay_min"))
            if c.has_option("WARMUP", "delay_max"):
                self.warmup_dmax.set(c.get("WARMUP", "delay_max"))

        if c.has_section("OUTREACH"):
            if c.has_option("OUTREACH", "subject"):
                self.email_subject.delete(0, tk.END)
                self.email_subject.insert(0, c.get("OUTREACH", "subject"))
            if c.has_option("OUTREACH", "body"):
                self.email_body.delete("1.0", tk.END)
                self.email_body.insert("1.0", c.get("OUTREACH", "body"))
            if c.has_option("OUTREACH", "daily_limit"):
                self.daily_limit.set(c.get("OUTREACH", "daily_limit"))
            if c.has_option("OUTREACH", "delay_min"):
                self.out_dmin.set(c.get("OUTREACH", "delay_min"))
            if c.has_option("OUTREACH", "delay_max"):
                self.out_dmax.set(c.get("OUTREACH", "delay_max"))

        if c.has_section("MONITOR") and c.has_option("MONITOR", "check_interval"):
            self.check_interval.set(c.get("MONITOR", "check_interval"))

    def _get_outreach_working_path(self):
        idx = self.outreach_import.current()
        if idx < 0 or idx >= len(self._outreach_paths):
            return ""
        return self._outreach_paths[idx]

    def _refresh_import_lists(self, select_working=""):
        preserve = (select_working or self._get_outreach_working_path() or "").strip()
        sel = self.db_imports_lb.curselection()
        preserve_iid = self._db_list_import_ids[sel[0]] if sel and self._db_list_import_ids else None
        imports = list_imports()
        for r in imports:
            try:
                resync_import_from_working(r["id"])
            except Exception:
                pass
        imports = list_imports()
        self._db_list_import_ids = [r["id"] for r in imports]
        self.db_imports_lb.delete(0, tk.END)
        for r in imports:
            self.db_imports_lb.insert(
                tk.END,
                f"#{r['id']} — {r['label']}  |  {r['imported_at'][:19]}",
            )
        labels = []
        paths = []
        for r in imports:
            labels.append(f"#{r['id']} — {r['label']}")
            paths.append(r["working_path"])
        self._outreach_paths = paths
        self.outreach_import["values"] = tuple(labels)
        if paths:
            pick = -1
            if preserve:
                ap = os.path.abspath(preserve)
                for i, p in enumerate(paths):
                    if os.path.abspath(p) == ap:
                        pick = i
                        break
            if pick < 0:
                pick = 0
            self.outreach_import.current(pick)
        else:
            self.outreach_import.set("")
        if self.db_imports_lb.size() > 0:
            pick = 0
            if preserve_iid is not None:
                for j, xid in enumerate(self._db_list_import_ids):
                    if xid == preserve_iid:
                        pick = j
                        break
            self.db_imports_lb.selection_set(pick)
            self.db_imports_lb.see(pick)
        self._on_database_list_select()

    def _import_database_dialog(self):
        p = filedialog.askopenfilename(filetypes=[("SQLite DB", "*.db"), ("All files", "*.*")])
        if not p:
            return
        try:
            iid, wp = import_user_database(p)
            self.log(f"📥 Imported database → working copy id {iid} (original left unchanged).")
            self._refresh_import_lists(select_working=wp)
            self._save_all_ui_prefs()
            messagebox.showinfo(
                "Imported",
                "A working copy was saved in the tool data folder.\n"
                "Your original file was not modified.\n\n"
                f"Import id: {iid}",
            )
        except PermissionError as ex:
            messagebox.showerror(
                "Import failed (permission denied)",
                f"{ex}\n\nTry moving the source .db to a normal folder you own, "
                "or run the app from a writable location.",
            )
        except Exception as ex:
            messagebox.showerror("Import failed", str(ex))

    def _import_excel_dialog(self):
        p = filedialog.askopenfilename(
            filetypes=[
                ("Excel .xlsx", "*.xlsx"),
                ("Excel .xlsm", "*.xlsm"),
                ("All files", "*.*"),
            ]
        )
        if not p:
            return
        try:
            iid, wp = import_excel_as_leads(p)
            try:
                st = get_stats_for_import(iid)
                n = st.get("total_rows", 0)
            except Exception:
                n = 0
            self.log(f"📊 Imported Excel as working database id {iid} (contacts rows: {n}).")
            self._refresh_import_lists(select_working=wp)
            self._save_all_ui_prefs()
            messagebox.showinfo(
                "Excel imported",
                "The spreadsheet was converted to SQLite and added to your lists.\n"
                "Your original Excel file was not modified.\n\n"
                f"Import id: {iid}\n\n"
                "Tip: row 1 should name columns (e.g. Full name, Email, Number, Sent, Replied).",
            )
        except PermissionError as ex:
            messagebox.showerror(
                "Import failed (permission denied)",
                str(ex),
            )
        except Exception as ex:
            messagebox.showerror("Excel import failed", str(ex))

    def _remove_selected_import(self):
        sel = self.db_imports_lb.curselection()
        if not sel:
            messagebox.showwarning("Remove import", "Select a list in the box first.")
            return
        iid = self._db_list_import_ids[sel[0]]
        wp = get_working_path_for_import(iid)
        if not messagebox.askyesno(
            "Remove import",
            "Remove this list from the tool and delete its working copy?\n\n"
            "Your original source .db file is not changed.",
        ):
            return
        prev_outreach = self._get_outreach_working_path()
        try:
            remove_import(iid)
        except Exception as ex:
            messagebox.showerror("Remove failed", str(ex))
            return
        self.log(f"🗑 Removed import #{iid} from tool.")
        next_sel = "" if (prev_outreach and wp and os.path.abspath(prev_outreach) == os.path.abspath(wp)) else prev_outreach
        self._refresh_import_lists(select_working=next_sel)
        self._save_all_ui_prefs()

    def _on_database_list_select(self):
        sel = self.db_imports_lb.curselection()
        if not sel:
            self.db_stat_total.config(text="Leads (with email): —")
            self.db_stat_sent.config(text="Sent: —")
            self.db_stat_replied.config(text="Replied: —")
            self.db_stat_left.config(text="Left (not sent): —")
            self.db_stat_note.config(text="")
            self._clear_db_location_breakdown()
            return
        iid = self._db_list_import_ids[sel[0]]
        st = get_stats_for_import(iid)
        self.db_stat_total.config(text=f"Leads (with email): {st['total']}")
        self.db_stat_sent.config(text=f"Sent: {st['sent']}")
        self.db_stat_replied.config(text=f"Replied: {st['replied']}")
        self.db_stat_left.config(text=f"Left (not sent): {st['left']}")
        self.db_stat_note.config(
            text=(
                f"All rows in table contacts: {st['total_rows']} · "
                f"Skipped (NULL / blank email): {st['no_email']}"
            )
        )
        self._refresh_db_location_breakdown(iid)

    def _add_lead_from_database_tab(self):
        sel = self.db_imports_lb.curselection()
        if not sel:
            messagebox.showerror("Error", "Select an imported list in the list above.")
            return
        keep_import_id = self._db_list_import_ids[sel[0]]
        try:
            add_lead_to_import(
                keep_import_id,
                self.db_add_name.get().strip(),
                self.db_add_phone.get().strip(),
                self.db_add_email.get().strip(),
            )
        except Exception as ex:
            messagebox.showerror("Could not add lead", str(ex))
            return
        self.db_add_name.delete(0, tk.END)
        self.db_add_phone.delete(0, tk.END)
        self.db_add_email.delete(0, tk.END)
        wp = self._get_outreach_working_path()
        self._refresh_import_lists(select_working=wp)
        for idx, iid in enumerate(self._db_list_import_ids):
            if iid == keep_import_id:
                self.db_imports_lb.selection_set(idx)
                self.db_imports_lb.see(idx)
                break
        self._on_database_list_select()
        messagebox.showinfo("Added", "Lead added to the working copy and master database.")

    def _start_warmup(self):
        emails = [e for e in self.warmup_emails.get("1.0","end").strip().splitlines() if e.strip()]
        if not emails:
            messagebox.showerror("Error","Enter at least one warmup email address.")
            return
        self.stop_warmup.clear()
        t = threading.Thread(target=run_warmup, args=(
            emails,
            int(self.warmup_count.get()),
            int(self.warmup_dmin.get()),
            int(self.warmup_dmax.get()),
            self.log,
            self.stop_warmup
        ), daemon=True)
        t.start()

    def _stop_warmup(self):
        self.stop_warmup.set()

    def _apply_outreach_indicator(self, running: bool):
        """Main thread only. Updates header + bottom bar; keeps busy flag in sync."""
        self._outreach_busy = running
        if running:
            text = "Outreach: RUNNING"
            fg = "#FFB86C"
        else:
            text = "Outreach: Idle"
            fg = "#666666"
        self.hdr_outreach_status.config(text=text, fg=fg)
        self.status_outreach.config(text=text, fg=fg)

    def _set_outreach_countdown(self, seconds_left: int):
        """Main thread only. Updates OUTREACH label with seconds until next send."""
        if not self._outreach_busy:
            return
        try:
            secs = int(seconds_left)
        except Exception:
            secs = 0
        if secs > 0:
            text = f"Outreach: RUNNING ({secs}s)"
        else:
            text = "Outreach: RUNNING"
        fg = "#FFB86C"
        self.hdr_outreach_status.config(text=text, fg=fg)
        self.status_outreach.config(text=text, fg=fg)

    def _start_outreach(self):
        db = self._get_outreach_working_path()
        if not db or not os.path.exists(db):
            messagebox.showerror(
                "Error",
                "Import a .db from the Database tab (or use Import .db…) and select a list.",
            )
            return
        if self._outreach_busy:
            messagebox.showwarning(
                "Outreach",
                "Outreach is already running. Use Stop or wait until it finishes.",
            )
            return
        thr = getattr(self, "_outreach_thread", None)
        if thr is not None and thr.is_alive():
            messagebox.showwarning(
                "Outreach",
                "Outreach is still finishing. Wait a moment before starting again.",
            )
            return
        self.stop_outreach.clear()
        self._apply_outreach_indicator(True)

        def worker():
            try:
                run_outreach(
                    db,
                    self.email_subject.get(),
                    self.email_body.get("1.0", "end").strip(),
                    int(self.daily_limit.get()),
                    int(self.out_dmin.get()),
                    int(self.out_dmax.get()),
                    self.log,
                    self.stop_outreach,
                    status_fn=lambda secs: self.after(0, lambda: self._set_outreach_countdown(secs)),
                )
            finally:
                self.after(0, lambda: self._apply_outreach_indicator(False))

        self._outreach_thread = threading.Thread(target=worker, daemon=True)
        self._outreach_thread.start()

    def _stop_outreach(self):
        self.stop_outreach.set()
        if self._outreach_busy:
            self.log("⏹ Stop requested — outreach will halt after the current step (send or wait).")

    def _apply_monitor_indicator(self, running: bool):
        """Main thread only. Header + bottom bar for reply monitor."""
        self._monitor_busy = running
        if running:
            text = "Reply monitor: RUNNING"
            fg = "#88C0FF"
        else:
            text = "Reply monitor: Idle"
            fg = "#666666"
        self.hdr_monitor_status.config(text=text, fg=fg)
        self.status_monitor.config(text=text, fg=fg)

    def _start_monitor(self):
        webhook = cfg("DISCORD","webhook")
        if not webhook:
            messagebox.showerror("Error","Set your Discord webhook URL in the Settings tab first.")
            return
        if self._monitor_busy:
            messagebox.showwarning(
                "Reply monitor",
                "The monitor is already running. Use Stop or wait until it finishes.",
            )
            return
        mthr = getattr(self, "_monitor_thread", None)
        if mthr is not None and mthr.is_alive():
            messagebox.showwarning(
                "Reply monitor",
                "The monitor is still finishing. Wait a moment before starting again.",
            )
            return
        self.stop_monitor.clear()
        self._apply_monitor_indicator(True)
        interval = int(self.check_interval.get())

        def worker():
            try:
                run_reply_monitor(webhook, interval, self.log, self.stop_monitor, self.seen_ids)
            finally:
                self.after(0, lambda: self._apply_monitor_indicator(False))

        self._monitor_thread = threading.Thread(target=worker, daemon=True)
        self._monitor_thread.start()

    def _stop_monitor(self):
        self.stop_monitor.set()
        if self._monitor_busy:
            self.log("⏹ Stop requested — reply monitor will stop after the current inbox pass / wait.")

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    app = App()
    app.mainloop()
